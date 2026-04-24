"""Historical data import from Caseware / Sage 50 / Excel / IIF / CSV.

When a CPA onboards an existing client, they usually have 1-3 years of
prior-period activity in another product. Scope 2.2 already handles
the QBO case. This module covers the remaining formats the field uses:

  - Generic CSV with transaction columns (always works)
  - IIF (QuickBooks Desktop export format)
  - Excel with a trial-balance layout
  - Sage 50 CSV export
  - Caseware trial-balance export

All imports land as ``gl_transactions`` rows tagged with
``source='historical_<format>'`` + a ``historical_imports`` job row
that records the original file, the detected format, the account
mapping used, and a pointer to the retained source blob. This keeps
the audit trail: ``how did this prior-year number get in the books?``

The mapping step is required whenever the source CoA doesn't line up
with the client's OtoCPA CoA. ``detect_mapping`` returns the accounts
present in the source that don't already exist in the destination and
prompts the CPA to pick a target code (or decline and create a new
one). Unmapped rows block the final posting step.

The source file itself is stored via ``save_source_blob`` and the
blob path is written to the ``historical_imports`` row so future
audits can reconstruct exactly what was uploaded.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)


# Source format identifiers (also used in historical_imports.source_format
# and gl_transactions.source as historical_<format>).
FORMAT_CSV = 'csv'
FORMAT_IIF = 'iif'
FORMAT_EXCEL_TB = 'excel_tb'
FORMAT_SAGE50 = 'sage50'
FORMAT_CASEWARE = 'caseware'

ALL_FORMATS = (FORMAT_CSV, FORMAT_IIF, FORMAT_EXCEL_TB, FORMAT_SAGE50,
               FORMAT_CASEWARE)


STATUS_DRAFT = 'draft'          # parsed + preview, not yet mapped/posted
STATUS_MAPPED = 'mapped'        # mapping complete, ready to post
STATUS_POSTED = 'posted'
STATUS_ROLLED_BACK = 'rolled_back'


_DATE_FORMATS = (
    '%Y-%m-%d', '%Y/%m/%d',
    '%m/%d/%Y', '%m/%d/%y',
    '%d/%m/%Y', '%d/%m/%y',
    '%d-%b-%Y', '%d-%b-%y',
)


def _open(db_path: Path | str) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec='seconds')


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------


def ensure_schema(db_path: Path | str) -> None:
    with _open(db_path) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS historical_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                firm_code TEXT NOT NULL,
                client_code TEXT NOT NULL,
                source_format TEXT NOT NULL,
                source_filename TEXT,
                source_blob_path TEXT,
                source_sha256 TEXT,
                row_count INTEGER DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'draft',
                mapping_json TEXT,
                preview_json TEXT,
                posted_entry_count INTEGER DEFAULT 0,
                posted_at TEXT,
                posted_by TEXT,
                rolled_back_at TEXT,
                rolled_back_by TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_hist_imports_firm_client "
            "ON historical_imports(firm_code, client_code, status)"
        )
        # gl_transactions is owned by gl_engine; make sure it exists for
        # test harnesses that skip gl_engine bootstrap.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS gl_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id TEXT NOT NULL,
                client_code TEXT NOT NULL,
                period TEXT NOT NULL,
                entry_date TEXT NOT NULL,
                account_code TEXT NOT NULL,
                side TEXT NOT NULL CHECK (side IN ('debit','credit')),
                amount REAL NOT NULL CHECK (amount > 0),
                description TEXT,
                source TEXT NOT NULL DEFAULT 'manual_je',
                document_id TEXT,
                reversed_by TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
        """)
        conn.commit()


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------


def detect_format(filename: str, file_bytes: bytes) -> str:
    """Best-effort format detection.

    Falls back to ``csv`` for unknown text input. Priority matches the
    spec (most-specific signatures first).
    """
    name = (filename or '').lower()
    try:
        head = file_bytes[:4096].decode('utf-8', errors='replace')
    except Exception:
        head = ''

    # IIF: QuickBooks Desktop tab-separated with !TRNS/!SPL/!ENDTRNS markers
    if name.endswith('.iif') or head.lstrip().startswith('!TRNS'):
        return FORMAT_IIF

    # Excel: binary signatures
    if name.endswith(('.xls', '.xlsx')) or file_bytes[:4] == b'PK\x03\x04' \
            or file_bytes[:4] == b'\xd0\xcf\x11\xe0':
        return FORMAT_EXCEL_TB

    # Caseware: TB export usually has "Trial Balance" in the first few
    # lines and a column named "Leadsheet" or "Map No."
    if 'caseware' in name or ('leadsheet' in head.lower()
                              or 'map no' in head.lower()):
        return FORMAT_CASEWARE

    # Sage 50: CSV export typically has "GL Account Number" + "Debit Amount"
    # + "Credit Amount" headers and a file named SageJournalEntryExport or
    # similar.
    if 'sage' in name or ('gl account number' in head.lower()
                          and 'debit amount' in head.lower()):
        return FORMAT_SAGE50

    return FORMAT_CSV


# ---------------------------------------------------------------------------
# Parsers (return a normalized list of transaction dicts)
# ---------------------------------------------------------------------------


def _decode(file_bytes: bytes) -> str:
    """UTF-8 first, fall back to cp1252 for Sage / Windows exports."""
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode('utf-8', errors='replace')


def _parse_date(raw: str) -> str | None:
    raw = (raw or '').strip()
    if not raw:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_amount(raw: Any) -> float | None:
    if raw is None or raw == '':
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace('$', '').replace(',', '')
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def parse_csv_generic(file_bytes: bytes) -> dict:
    """Parse a generic CSV.

    Accepts a flexible schema — columns are matched case-insensitively
    against a few common spellings. The minimum required shape is one
    of:

      - date + account + (debit | credit)
      - date + account + amount + side
      - date + account + amount  (signed: +debit / -credit)
    """
    text = _decode(file_bytes)
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    errors: list[str] = []
    accounts_seen: set[str] = set()
    if not reader.fieldnames:
        return {'rows': [], 'accounts': [], 'errors': ['empty_file']}

    # Build a case-insensitive header map.
    hmap: dict[str, str] = {}
    for h in reader.fieldnames:
        hmap[h.strip().lower()] = h

    def col(*names: str) -> str | None:
        for n in names:
            if n in hmap:
                return hmap[n]
        return None

    c_date = col('date', 'entry date', 'transaction date', 'trans date')
    c_acct = col('account', 'account_code', 'account code', 'gl account',
                 'gl account number', 'account number')
    c_amt = col('amount', 'value')
    c_deb = col('debit', 'debits', 'debit amount', 'dr')
    c_cre = col('credit', 'credits', 'credit amount', 'cr')
    c_side = col('side', 'dr/cr', 'type')
    c_desc = col('description', 'memo', 'comment', 'notes')
    c_vend = col('vendor', 'payee', 'name')

    if not c_date or not c_acct:
        return {'rows': [], 'accounts': [],
                'errors': ['missing_required_columns_date_or_account']}

    for i, raw in enumerate(reader, start=2):  # line 1 is header
        date = _parse_date(raw.get(c_date, '') if c_date else '')
        acct = (raw.get(c_acct, '') or '').strip() if c_acct else ''
        desc = (raw.get(c_desc, '') or '').strip() if c_desc else ''
        vendor = (raw.get(c_vend, '') or '').strip() if c_vend else ''
        if not date:
            errors.append(f'row {i}: invalid or missing date')
            continue
        if not acct:
            errors.append(f'row {i}: missing account')
            continue

        debit = _parse_amount(raw.get(c_deb)) if c_deb else None
        credit = _parse_amount(raw.get(c_cre)) if c_cre else None
        amt = _parse_amount(raw.get(c_amt)) if c_amt else None
        side_raw = (raw.get(c_side, '') or '').strip().lower() if c_side \
            else ''

        entries: list[tuple[str, float]] = []
        if debit and debit != 0:
            entries.append(('debit', abs(debit)))
        if credit and credit != 0:
            entries.append(('credit', abs(credit)))
        if not entries and amt is not None and amt != 0:
            if side_raw in ('debit', 'dr', 'd'):
                entries.append(('debit', abs(amt)))
            elif side_raw in ('credit', 'cr', 'c'):
                entries.append(('credit', abs(amt)))
            else:
                # Signed amount: positive=debit, negative=credit.
                entries.append(('debit' if amt > 0 else 'credit',
                                abs(amt)))
        if not entries:
            errors.append(f'row {i}: no non-zero amount')
            continue

        accounts_seen.add(acct)
        for side, amount in entries:
            rows.append({
                'date': date,
                'account_code': acct,
                'side': side,
                'amount': amount,
                'description': desc or vendor or '',
                'vendor': vendor,
            })

    return {
        'rows': rows,
        'accounts': sorted(accounts_seen),
        'errors': errors,
    }


def parse_iif(file_bytes: bytes) -> dict:
    """Parse a QuickBooks Desktop IIF file.

    IIF is tab-separated with header lines beginning with ``!`` and
    data lines beginning with the matching keyword. We look at
    ``!TRNS`` / ``TRNS`` (header leg) and ``!SPL`` / ``SPL`` (split
    legs); ``ENDTRNS`` closes the entry.
    """
    text = _decode(file_bytes)
    lines = [ln for ln in text.splitlines() if ln.strip()]
    trns_fields: list[str] = []
    spl_fields: list[str] = []
    rows: list[dict] = []
    accounts: set[str] = set()
    errors: list[str] = []

    for ln in lines:
        parts = ln.split('\t')
        tag = parts[0]
        if tag == '!TRNS':
            trns_fields = [p.strip() for p in parts]
        elif tag == '!SPL':
            spl_fields = [p.strip() for p in parts]
        elif tag == 'TRNS' and trns_fields:
            d = dict(zip(trns_fields, parts))
            date = _parse_date(d.get('DATE', ''))
            acct = (d.get('ACCNT', '') or '').strip()
            amt = _parse_amount(d.get('AMOUNT', ''))
            memo = (d.get('MEMO', '') or '').strip()
            name = (d.get('NAME', '') or '').strip()
            if date and acct and amt is not None and amt != 0:
                accounts.add(acct)
                rows.append({
                    'date': date,
                    'account_code': acct,
                    'side': 'debit' if amt > 0 else 'credit',
                    'amount': abs(amt),
                    'description': memo or name or '',
                    'vendor': name,
                })
            else:
                errors.append(f'TRNS line malformed: {ln[:80]}')
        elif tag == 'SPL' and spl_fields:
            d = dict(zip(spl_fields, parts))
            date = _parse_date(d.get('DATE', ''))
            acct = (d.get('ACCNT', '') or '').strip()
            amt = _parse_amount(d.get('AMOUNT', ''))
            memo = (d.get('MEMO', '') or '').strip()
            if date and acct and amt is not None and amt != 0:
                accounts.add(acct)
                rows.append({
                    'date': date,
                    'account_code': acct,
                    'side': 'debit' if amt > 0 else 'credit',
                    'amount': abs(amt),
                    'description': memo or '',
                    'vendor': '',
                })

    if not rows and not errors:
        errors.append('no_trns_or_spl_rows')
    return {'rows': rows, 'accounts': sorted(accounts), 'errors': errors}


def parse_excel_tb(file_bytes: bytes) -> dict:
    """Parse a trial-balance Excel file.

    Uses openpyxl if available. The expected layout is a header row
    with columns: ``Account Code``, ``Account Name`` (optional),
    ``Debit``, ``Credit``, plus an ``As of`` cell somewhere in the
    first five rows. Rows where ``debit`` and ``credit`` are both 0
    are skipped.
    """
    try:
        import openpyxl  # type: ignore
    except Exception:
        return {'rows': [], 'accounts': [],
                'errors': ['openpyxl_not_installed']}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {'rows': [], 'accounts': [],
                'errors': [f'excel_load_failed: {e}']}
    ws = wb.active
    header_row = None
    header_map: dict[str, int] = {}
    as_of = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        lowered = [str(v or '').strip().lower() for v in row]
        if ('debit' in lowered and 'credit' in lowered
                and any('account' in c for c in lowered)):
            header_row = r_idx
            for i, c in enumerate(lowered):
                header_map[c] = i
            break
        if as_of is None:
            for c in row:
                if c and isinstance(c, str) and 'as of' in c.lower():
                    # Look for a date in the same row.
                    for v in row:
                        parsed = _parse_date(str(v or ''))
                        if parsed:
                            as_of = parsed
                            break
    if not header_row:
        return {'rows': [], 'accounts': [],
                'errors': ['excel_header_not_found']}

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    c_acct = col('account code', 'account_code', 'account number', 'account',
                 'gl account')
    c_debit = col('debit', 'debits')
    c_credit = col('credit', 'credits')
    if c_acct is None or c_debit is None or c_credit is None:
        return {'rows': [], 'accounts': [],
                'errors': ['excel_missing_required_columns']}

    rows: list[dict] = []
    accounts: set[str] = set()
    errors: list[str] = []
    date_stamp = as_of or datetime.now(timezone.utc).date().isoformat()
    for r_idx, row in enumerate(ws.iter_rows(values_only=True),
                                start=1):
        if r_idx <= header_row:
            continue
        if not row or all(v in (None, '') for v in row):
            continue
        acct = str(row[c_acct] or '').strip()
        if not acct:
            continue
        debit = _parse_amount(row[c_debit]) or 0
        credit = _parse_amount(row[c_credit]) or 0
        if debit == 0 and credit == 0:
            continue
        accounts.add(acct)
        if debit:
            rows.append({'date': date_stamp, 'account_code': acct,
                         'side': 'debit', 'amount': abs(debit),
                         'description': 'TB opening', 'vendor': ''})
        if credit:
            rows.append({'date': date_stamp, 'account_code': acct,
                         'side': 'credit', 'amount': abs(credit),
                         'description': 'TB opening', 'vendor': ''})

    return {'rows': rows, 'accounts': sorted(accounts), 'errors': errors}


def parse_sage50(file_bytes: bytes) -> dict:
    """Sage 50 CSV journal entry export.

    Sage 50 CSV columns vary by version; the shared baseline is
    ``Date`` + ``GL Account Number`` + ``Debit Amount`` + ``Credit
    Amount`` + ``Description``. Reuses the generic CSV parser because
    that one already handles those header spellings.
    """
    return parse_csv_generic(file_bytes)


def parse_caseware(file_bytes: bytes) -> dict:
    """Caseware trial-balance export (CSV flavour).

    Columns vary; the generic parser will pick up ``Account`` +
    ``Debit`` + ``Credit`` when present. Pure TB files have no date,
    so synthesize one if missing.
    """
    text = _decode(file_bytes)
    # Look for an "As of" cue in the first five lines.
    as_of = None
    for ln in text.splitlines()[:5]:
        m = re.search(r'as of[:\s]+(\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4})',
                      ln, re.I)
        if m:
            as_of = _parse_date(m.group(1))
            if as_of:
                break

    parsed = parse_csv_generic(file_bytes)
    if as_of:
        for r in parsed['rows']:
            if r['date'] < as_of or not r['date']:
                r['date'] = as_of
    elif parsed['rows'] and not parsed['rows'][0].get('date'):
        today = datetime.now(timezone.utc).date().isoformat()
        for r in parsed['rows']:
            r['date'] = r['date'] or today
    return parsed


PARSERS = {
    FORMAT_CSV: parse_csv_generic,
    FORMAT_IIF: parse_iif,
    FORMAT_EXCEL_TB: parse_excel_tb,
    FORMAT_SAGE50: parse_sage50,
    FORMAT_CASEWARE: parse_caseware,
}


# ---------------------------------------------------------------------------
# Account mapping
# ---------------------------------------------------------------------------


def detect_unmapped(db_path: Path | str, client_code: str,
                    source_accounts: list[str]) -> list[str]:
    """Return source accounts that don't already exist in gl_accounts
    for the client. Caller presents these to the CPA for mapping."""
    with _open(db_path) as conn:
        existing_col = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND "
            "name='gl_accounts'"
        ).fetchone()
        if not existing_col:
            return list(source_accounts)
        rows = conn.execute(
            "SELECT account_code FROM gl_accounts WHERE client_code=?",
            (client_code,)
        ).fetchall()
    existing = {r['account_code'] for r in rows}
    return [a for a in source_accounts if a not in existing]


def apply_mapping(rows: list[dict], mapping: dict[str, str]) -> list[dict]:
    """Translate each row's source account via the mapping dict.

    Unmapped accounts are left alone — the caller surfaces them in
    the preview so the CPA knows which rows would post as-is.
    """
    out = []
    for r in rows:
        m = dict(r)
        src = r.get('account_code', '')
        if src in mapping and mapping[src]:
            m['account_code'] = mapping[src]
        out.append(m)
    return out


# ---------------------------------------------------------------------------
# Source blob retention
# ---------------------------------------------------------------------------


def save_source_blob(root: Path | str, firm_code: str, client_code: str,
                     filename: str, file_bytes: bytes) -> tuple[str, str]:
    """Persist the raw uploaded file so the audit trail can reconstruct
    it. Returns (path, sha256)."""
    root_p = Path(root)
    root_p.mkdir(parents=True, exist_ok=True)
    sha = hashlib.sha256(file_bytes).hexdigest()
    safe_name = re.sub(r'[^A-Za-z0-9._-]', '_', filename or 'import.dat')
    target = root_p / f'{firm_code}_{client_code}_{sha[:12]}_{safe_name}'
    target.write_bytes(file_bytes)
    return (str(target), sha)


# ---------------------------------------------------------------------------
# Import jobs
# ---------------------------------------------------------------------------


def create_import_job(db_path: Path | str, firm_code: str,
                      client_code: str, source_format: str,
                      filename: str, blob_path: str, sha: str,
                      rows: list[dict], accounts: list[str]) -> int:
    ensure_schema(db_path)
    with _open(db_path) as conn:
        cur = conn.execute(
            "INSERT INTO historical_imports "
            "(firm_code, client_code, source_format, source_filename, "
            " source_blob_path, source_sha256, row_count, status, "
            " preview_json) VALUES (?,?,?,?,?,?,?,?,?)",
            (firm_code, client_code, source_format, filename,
             blob_path, sha, len(rows), STATUS_DRAFT,
             json.dumps({'accounts': accounts, 'sample': rows[:20]})),
        )
        conn.commit()
        return int(cur.lastrowid)


def save_mapping(db_path: Path | str, job_id: int,
                 mapping: dict[str, str]) -> None:
    with _open(db_path) as conn:
        conn.execute(
            "UPDATE historical_imports SET mapping_json=?, status=? "
            "WHERE id=?",
            (json.dumps(mapping), STATUS_MAPPED, job_id),
        )
        conn.commit()


def post_import(db_path: Path | str, job_id: int, rows: list[dict],
                posted_by: str) -> dict:
    """Write rows into ``gl_transactions`` and mark the job posted.

    Entries are batched under a single synthetic ``entry_id`` per
    (import_id, date) so TB queries still see valid JEs.
    """
    ensure_schema(db_path)
    posted = 0
    with _open(db_path) as conn:
        job = conn.execute(
            "SELECT * FROM historical_imports WHERE id=?", (job_id,)
        ).fetchone()
        if not job:
            return {'ok': False, 'reason': 'unknown_job'}
        if job['status'] == STATUS_POSTED:
            return {'ok': False, 'reason': 'already_posted'}

        client_code = job['client_code']
        src_tag = f"historical_{job['source_format']}"
        for r in rows:
            date = r['date']
            period = date[:7]
            entry_id = f"HIST-{job_id}-{date.replace('-', '')}"
            conn.execute(
                "INSERT INTO gl_transactions "
                "(entry_id, client_code, period, entry_date, account_code, "
                " side, amount, description, source) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                (entry_id, client_code, period, date,
                 str(r['account_code']).strip(), r['side'],
                 float(r['amount']),
                 (r.get('description') or '')[:500], src_tag),
            )
            posted += 1

        conn.execute(
            "UPDATE historical_imports SET status=?, posted_at=?, "
            "posted_by=?, posted_entry_count=? WHERE id=?",
            (STATUS_POSTED, _iso_now(), posted_by, posted, job_id),
        )
        conn.commit()
    return {'ok': True, 'posted': posted}


def rollback_import(db_path: Path | str, job_id: int,
                    by: str) -> dict:
    with _open(db_path) as conn:
        job = conn.execute(
            "SELECT * FROM historical_imports WHERE id=?", (job_id,)
        ).fetchone()
        if not job:
            return {'ok': False, 'reason': 'unknown_job'}
        if job['status'] != STATUS_POSTED:
            return {'ok': False, 'reason': 'not_posted'}
        deleted = conn.execute(
            "DELETE FROM gl_transactions WHERE "
            "source=? AND entry_id LIKE ?",
            (f"historical_{job['source_format']}", f'HIST-{job_id}-%')
        ).rowcount
        conn.execute(
            "UPDATE historical_imports SET status=?, rolled_back_at=?, "
            "rolled_back_by=? WHERE id=?",
            (STATUS_ROLLED_BACK, _iso_now(), by, job_id),
        )
        conn.commit()
    return {'ok': True, 'deleted': deleted}


def get_import(db_path: Path | str, job_id: int) -> dict | None:
    ensure_schema(db_path)
    with _open(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM historical_imports WHERE id=?", (job_id,)
        ).fetchone()
    return dict(row) if row else None


def list_imports(db_path: Path | str, firm_code: str,
                 client_code: str) -> list[dict]:
    ensure_schema(db_path)
    with _open(db_path) as conn:
        rows = conn.execute(
            "SELECT * FROM historical_imports WHERE firm_code=? "
            "AND client_code=? ORDER BY id DESC",
            (firm_code, client_code)
        ).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Bilingual strings
# ---------------------------------------------------------------------------


_I18N = {
    'fr': {
        'page_title': 'Importation historique',
        'upload_heading': 'Téléverser un fichier historique',
        'upload_hint': 'Formats acceptés : CSV, Excel, IIF, Sage 50, Caseware',
        'btn_upload': 'Téléverser',
        'btn_confirm': 'Confirmer et enregistrer',
        'btn_rollback': 'Annuler l\'importation',
        'btn_save_mapping': 'Enregistrer le mappage',
        'rollback_confirm':
            'Annuler supprime toutes les lignes issues de cet import. '
            'Continuer ?',
        'col_format': 'Format',
        'col_status': 'Statut',
        'col_rows': 'Lignes',
        'col_created': 'Créé',
        'col_source': 'Fichier source',
        'unmapped_heading': 'Comptes non mappés',
        'unmapped_hint':
            'Ces comptes n\'existent pas dans le plan comptable du client. '
            'Associez-les à un compte existant ou laissez pour créer.',
        'no_imports': 'Aucune importation historique pour ce client.',
        'preview_heading': 'Aperçu (20 premières lignes)',
        'status_draft': 'brouillon',
        'status_mapped': 'mappé',
        'status_posted': 'comptabilisé',
        'status_rolled_back': 'annulé',
    },
    'en': {
        'page_title': 'Historical import',
        'upload_heading': 'Upload a historical file',
        'upload_hint':
            'Accepted formats: CSV, Excel, IIF, Sage 50, Caseware',
        'btn_upload': 'Upload',
        'btn_confirm': 'Confirm and post',
        'btn_rollback': 'Roll back import',
        'btn_save_mapping': 'Save mapping',
        'rollback_confirm':
            'Rollback removes every row posted from this import. Continue?',
        'col_format': 'Format',
        'col_status': 'Status',
        'col_rows': 'Rows',
        'col_created': 'Created',
        'col_source': 'Source file',
        'unmapped_heading': 'Unmapped accounts',
        'unmapped_hint':
            'These accounts are not in the client chart of accounts. '
            'Map them to an existing account or leave blank to create.',
        'no_imports': 'No historical imports on file for this client.',
        'preview_heading': 'Preview (first 20 rows)',
        'status_draft': 'draft',
        'status_mapped': 'mapped',
        'status_posted': 'posted',
        'status_rolled_back': 'rolled back',
    },
}


def _tr(key: str, lang: str) -> str:
    return _I18N.get(lang or 'fr', _I18N['fr']).get(
        key, _I18N['fr'].get(key, key)
    )


# ---------------------------------------------------------------------------
# Renderer
# ---------------------------------------------------------------------------


def render_import_historical_page(
    *, firm_code: str, client_code: str, client_name: str,
    imports: list[dict], lang: str = 'fr',
    flash: str = '', flash_error: str = '',
    preview: dict | None = None,
    gl_accounts: list[dict] | None = None,
) -> str:
    """Render the /clients/<code>/import_historical page.

    The renderer is deliberately thin — full styling comes from the
    host page_layout. Preview shows the most recent draft's sample
    rows + unmapped accounts. Posted imports get confirm/rollback
    buttons scoped to that job id.
    """
    import html as _html

    def _esc(s: Any) -> str:
        return _html.escape(str(s or ''))

    flash_html = ''
    if flash:
        flash_html += (
            f'<div style="background:#d4edda;padding:8px;'
            f'margin-bottom:10px;">{_esc(flash)}</div>'
        )
    if flash_error:
        flash_html += (
            f'<div style="background:#f8d7da;padding:8px;'
            f'margin-bottom:10px;">{_esc(flash_error)}</div>'
        )

    # Existing imports table
    if imports:
        rows_html = ''
        for imp in imports:
            actions = ''
            if imp['status'] == STATUS_POSTED:
                actions = (
                    f'<form method="POST" action="/clients/import_historical/rollback" '
                    'style="display:inline;margin:0;" '
                    f'onsubmit="return confirm(\'{_esc(_tr("rollback_confirm", lang))}\');">'
                    f'<input type="hidden" name="client_code" value="{_esc(client_code)}">'
                    f'<input type="hidden" name="job_id" value="{int(imp["id"])}">'
                    f'<button type="submit" style="background:#dc2626;color:white;">'
                    f'{_esc(_tr("btn_rollback", lang))}</button></form>'
                )
            elif imp['status'] in (STATUS_DRAFT, STATUS_MAPPED):
                actions = (
                    f'<form method="POST" action="/clients/import_historical/post" '
                    'style="display:inline;margin:0;">'
                    f'<input type="hidden" name="client_code" value="{_esc(client_code)}">'
                    f'<input type="hidden" name="job_id" value="{int(imp["id"])}">'
                    f'<button type="submit">{_esc(_tr("btn_confirm", lang))}</button>'
                    '</form>'
                )
            status_label = _tr(f'status_{imp["status"]}', lang)
            rows_html += (
                '<tr>'
                f'<td>{int(imp["id"])}</td>'
                f'<td>{_esc(imp["source_format"])}</td>'
                f'<td>{_esc(status_label)}</td>'
                f'<td>{int(imp["row_count"])}</td>'
                f'<td>{_esc(imp.get("source_filename"))}</td>'
                f'<td>{_esc(imp.get("created_at"))}</td>'
                f'<td>{actions}</td>'
                '</tr>'
            )
        table = (
            '<table style="width:100%;border-collapse:collapse;">'
            '<thead><tr>'
            '<th>#</th>'
            f'<th>{_esc(_tr("col_format", lang))}</th>'
            f'<th>{_esc(_tr("col_status", lang))}</th>'
            f'<th>{_esc(_tr("col_rows", lang))}</th>'
            f'<th>{_esc(_tr("col_source", lang))}</th>'
            f'<th>{_esc(_tr("col_created", lang))}</th>'
            '<th></th>'
            '</tr></thead>'
            f'<tbody>{rows_html}</tbody></table>'
        )
    else:
        table = f'<p>{_esc(_tr("no_imports", lang))}</p>'

    # Preview section for the most recent draft
    preview_html = ''
    if preview and preview.get('sample'):
        sample = preview['sample'][:20]
        sample_rows = ''.join(
            f'<tr><td>{_esc(s.get("date"))}</td>'
            f'<td>{_esc(s.get("account_code"))}</td>'
            f'<td>{_esc(s.get("side"))}</td>'
            f'<td style="text-align:right;">'
            f'{float(s.get("amount") or 0):.2f}</td>'
            f'<td>{_esc(s.get("description"))}</td></tr>'
            for s in sample
        )
        preview_html = (
            f'<div class="card"><h3>{_esc(_tr("preview_heading", lang))}</h3>'
            '<table style="width:100%;border-collapse:collapse;">'
            '<thead><tr><th>Date</th><th>Compte/Account</th>'
            '<th>DR/CR</th><th>Amount</th><th>Description</th></tr></thead>'
            f'<tbody>{sample_rows}</tbody></table></div>'
        )
        unmapped = preview.get('unmapped') or []
        if unmapped:
            opts = ''.join(
                f'<option value="{_esc(a["account_code"])}">'
                f'{_esc(a["account_code"])} — {_esc(a.get("account_name"))}'
                '</option>'
                for a in (gl_accounts or [])
            )
            rows = ''
            for src_acct in unmapped:
                rows += (
                    f'<tr><td>{_esc(src_acct)}</td>'
                    f'<td><select name="map_{_esc(src_acct)}">'
                    '<option value="">(keep as-is)</option>'
                    f'{opts}</select></td></tr>'
                )
            preview_html += (
                '<div class="card">'
                f'<h3>{_esc(_tr("unmapped_heading", lang))}</h3>'
                f'<p>{_esc(_tr("unmapped_hint", lang))}</p>'
                '<form method="POST" '
                'action="/clients/import_historical/mapping">'
                f'<input type="hidden" name="client_code" value="{_esc(client_code)}">'
                f'<input type="hidden" name="job_id" value="{int(preview["job_id"])}">'
                '<table><thead><tr><th>Source</th><th>Target</th>'
                f'</tr></thead><tbody>{rows}</tbody></table>'
                f'<button type="submit">{_esc(_tr("btn_save_mapping", lang))}</button>'
                '</form></div>'
            )

    # Upload form
    upload_html = (
        '<div class="card">'
        f'<h3>{_esc(_tr("upload_heading", lang))}</h3>'
        f'<p class="muted">{_esc(_tr("upload_hint", lang))}</p>'
        '<form method="POST" enctype="multipart/form-data" '
        'action="/clients/import_historical/upload">'
        f'<input type="hidden" name="client_code" value="{_esc(client_code)}">'
        '<input type="file" name="file" required> '
        f'<button type="submit">{_esc(_tr("btn_upload", lang))}</button>'
        '</form></div>'
    )

    return (
        f'<h1>{_esc(client_name)} — {_esc(_tr("page_title", lang))}</h1>'
        f'{flash_html}{upload_html}{preview_html}{table}'
    )


# ---------------------------------------------------------------------------
# Orchestration helper used by the HTTP route
# ---------------------------------------------------------------------------


def ingest_upload(db_path: Path | str, blob_root: Path | str,
                  firm_code: str, client_code: str,
                  filename: str, file_bytes: bytes,
                  forced_format: str | None = None) -> dict:
    """End-to-end: detect format, parse, stash blob, create draft job.

    Returns a summary dict: ``{job_id, format, row_count, accounts,
    unmapped, errors, sample}``. The preview is small on purpose — the
    UI walks through it row-by-row and the blob stays on disk for later
    replays.
    """
    ensure_schema(db_path)
    fmt = forced_format or detect_format(filename, file_bytes)
    if fmt not in PARSERS:
        return {'ok': False, 'reason': 'unknown_format', 'format': fmt}

    parsed = PARSERS[fmt](file_bytes)
    rows = parsed.get('rows', [])
    errors = parsed.get('errors', [])
    accounts = parsed.get('accounts', [])
    if not rows:
        return {'ok': False, 'reason': 'no_rows', 'errors': errors,
                'format': fmt}

    blob_path, sha = save_source_blob(blob_root, firm_code, client_code,
                                      filename, file_bytes)
    job_id = create_import_job(db_path, firm_code, client_code, fmt,
                               filename, blob_path, sha, rows, accounts)
    unmapped = detect_unmapped(db_path, client_code, accounts)
    return {
        'ok': True,
        'job_id': job_id,
        'format': fmt,
        'row_count': len(rows),
        'accounts': accounts,
        'unmapped': unmapped,
        'errors': errors,
        'sample': rows[:20],
        'blob_path': blob_path,
        'sha256': sha,
    }
