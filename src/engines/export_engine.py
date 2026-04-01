"""
src/engines/export_engine.py — Multi-format accounting export engine.

Generates export files for:
- CSV (universal, UTF-8 BOM for Excel)
- Sage 50 Canada
- Acomba (Quebec, tab-delimited)
- QuickBooks Desktop (IIF)
- Xero
- Wave Accounting
- Excel (multi-sheet, openpyxl)
- Bulk ZIP (annual)

All functions accept a list of document dicts (from the DB) and return bytes.
No AI calls. Pure deterministic formatting.
"""
from __future__ import annotations

import calendar
import csv
import io
import sqlite3
import zipfile
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

ROOT_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = ROOT_DIR / "data" / "otocpa_agent.db"

CENT = Decimal("0.01")
_ZERO = Decimal("0")
GST_RATE = Decimal("0.05")
QST_RATE = Decimal("0.09975")

_FORMULA_PREFIXES = ("=", "+", "-", "@")


def sanitize_csv_cell(value: Any) -> str:
    """Neutralize CSV formula injection by prepending a quote to dangerous values."""
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in _FORMULA_PREFIXES:
        return "'" + s
    return s


def _round(value: Decimal) -> Decimal:
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def _dec(value: Any) -> Decimal:
    if value is None or str(value).strip() == "":
        return _ZERO
    try:
        return Decimal(str(value))
    except Exception:
        return _ZERO


def _extract_taxes(amount: Decimal, tax_code: str) -> dict[str, Decimal]:
    """Extract GST/QST/HST from a tax-inclusive amount based on tax code."""
    tc = (tax_code or "").strip().upper()
    if tc in ("T", "GST_QST"):
        divisor = Decimal("1") + GST_RATE + QST_RATE
        pre_tax = _round(amount / divisor)
        gst = _round(pre_tax * GST_RATE)
        qst = _round(pre_tax * QST_RATE)
        return {"pre_tax": pre_tax, "gst": gst, "qst": qst, "hst": _ZERO}
    elif tc == "M":
        divisor = Decimal("1") + GST_RATE + QST_RATE
        pre_tax = _round(amount / divisor)
        gst = _round(pre_tax * GST_RATE)
        qst = _round(pre_tax * QST_RATE)
        return {"pre_tax": pre_tax, "gst": gst, "qst": qst, "hst": _ZERO}
    elif tc == "HST":
        divisor = Decimal("1") + Decimal("0.13")
        pre_tax = _round(amount / divisor)
        hst = _round(pre_tax * Decimal("0.13"))
        return {"pre_tax": pre_tax, "gst": _ZERO, "qst": _ZERO, "hst": hst}
    elif tc == "HST_ATL":
        divisor = Decimal("1") + Decimal("0.15")
        pre_tax = _round(amount / divisor)
        hst = _round(pre_tax * Decimal("0.15"))
        return {"pre_tax": pre_tax, "gst": _ZERO, "qst": _ZERO, "hst": hst}
    else:
        return {"pre_tax": amount, "gst": _ZERO, "qst": _ZERO, "hst": _ZERO}


# ---------------------------------------------------------------------------
# DB query — shared by all exporters
# ---------------------------------------------------------------------------

def fetch_posted_documents(
    client_code: str,
    period_start: str,
    period_end: str,
    db_path: Path = DB_PATH,
) -> list[dict[str, Any]]:
    """Fetch posted documents for a client and date range."""
    if not db_path.exists():
        return []
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT
                d.document_id, d.vendor, d.document_date, d.amount,
                d.gl_account, d.tax_code, d.category, d.doc_type,
                d.file_name, d.client_code,
                COALESCE(pj.posting_status, '') AS posting_status,
                COALESCE(pj.external_id, '') AS external_id
            FROM documents d
            LEFT JOIN posting_jobs pj
                ON pj.document_id = d.document_id
                AND pj.rowid = (
                    SELECT pj2.rowid FROM posting_jobs pj2
                    WHERE pj2.document_id = d.document_id
                    ORDER BY COALESCE(pj2.updated_at, pj2.created_at) DESC,
                             pj2.rowid DESC LIMIT 1
                )
            WHERE LOWER(COALESCE(d.client_code, '')) = LOWER(?)
              AND COALESCE(d.document_date, '') >= ?
              AND COALESCE(d.document_date, '') <= ?
              AND (pj.posting_status = 'posted' OR COALESCE(pj.external_id, '') != '')
            ORDER BY d.document_date, d.document_id
            """,
            (client_code.strip(), period_start, period_end),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _period_dates(period: str) -> tuple[str, str]:
    """Convert '2026-01' to ('2026-01-01', '2026-01-31')."""
    parts = period.strip().split("-")
    year = int(parts[0])
    month = int(parts[1])
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


# ---------------------------------------------------------------------------
# PART 1 — CSV Universal Export
# ---------------------------------------------------------------------------

def generate_csv(docs: list[dict[str, Any]], **_kw: Any) -> bytes:
    """Generate a UTF-8 BOM CSV with standard columns."""
    buf = io.StringIO()
    buf.write("")  # We'll prepend BOM at byte level
    writer = csv.writer(buf)
    writer.writerow([
        "Date", "Vendor", "Description", "GL Account", "Amount",
        "GST", "QST", "HST", "Tax Code", "Document ID",
    ])
    for doc in docs:
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        writer.writerow([
            sanitize_csv_cell(doc.get("document_date", "")),
            sanitize_csv_cell(doc.get("vendor", "")),
            sanitize_csv_cell(f"{doc.get('vendor', '')} expense"),
            sanitize_csv_cell(doc.get("gl_account", "")),
            str(amount),
            str(taxes["gst"]),
            str(taxes["qst"]),
            str(taxes["hst"]),
            sanitize_csv_cell(doc.get("tax_code", "")),
            sanitize_csv_cell(doc.get("document_id", "")),
        ])
    # UTF-8 BOM for Excel French character support
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PART 2 — Sage 50 Canada Export
# ---------------------------------------------------------------------------

_SAGE50_TAX_MAP = {
    "T": "GP",
    "GST_QST": "GP",
    "E": "E",
    "M": "GP",
    "Z": "Z",
    "HST": "H",
    "HST_ATL": "H",
    "I": "E",
    "NONE": "",
    "": "",
}


def generate_sage50(docs: list[dict[str, Any]]) -> bytes:
    """Generate Sage 50 Canada import CSV."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Date", "Reference", "Description", "Account Number",
        "Debit", "Credit", "Tax Code",
    ])
    for doc in docs:
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        # Sage 50 date format: MM/DD/YYYY
        raw_date = doc.get("document_date", "")
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            sage_date = dt.strftime("%m/%d/%Y")
        except Exception:
            sage_date = raw_date
        tc = (doc.get("tax_code", "") or "").strip().upper()
        sage_tax = _SAGE50_TAX_MAP.get(tc, "")
        debit = str(taxes["pre_tax"])
        credit = ""
        writer.writerow([
            sanitize_csv_cell(sage_date),
            sanitize_csv_cell(doc.get("document_id", "")),
            sanitize_csv_cell(f"{doc.get('vendor', '')} expense"),
            sanitize_csv_cell(doc.get("gl_account", "")),
            debit,
            credit,
            sanitize_csv_cell(sage_tax),
        ])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PART 3 — Acomba Export (Quebec, tab-delimited)
# ---------------------------------------------------------------------------

def generate_acomba(docs: list[dict[str, Any]]) -> bytes:
    """Generate Acomba import file (tab-delimited, French headers)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="\t")
    writer.writerow([
        "No_Pièce", "Date", "No_Compte", "Description",
        "Débit", "Crédit", "TPS", "TVQ",
    ])
    for doc in docs:
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        # Acomba date: YYYYMMDD
        raw_date = doc.get("document_date", "")
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            acomba_date = dt.strftime("%Y%m%d")
        except Exception:
            acomba_date = raw_date.replace("-", "")
        writer.writerow([
            sanitize_csv_cell(doc.get("document_id", "")),
            sanitize_csv_cell(acomba_date),
            sanitize_csv_cell(doc.get("gl_account", "")),
            sanitize_csv_cell(f"{doc.get('vendor', '')} expense"),
            str(taxes["pre_tax"]),
            "",
            str(taxes["gst"]),
            str(taxes["qst"]),
        ])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PART 4 — QuickBooks Desktop IIF Export
# ---------------------------------------------------------------------------

def generate_qbd_iif(docs: list[dict[str, Any]]) -> bytes:
    """Generate QuickBooks Desktop IIF (Intuit Interchange Format)."""
    lines: list[str] = []
    # Header definitions
    lines.append(
        "!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tDOCNUM\tMEMO\tCLEAR\tTOPRINT\tADDR1"
    )
    lines.append(
        "!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tAMOUNT\tDOCNUM\tMEMO\tCLEAR\tQNTY\tPRICE\tINVMEMO"
    )
    lines.append("!ENDTRNS")

    for i, doc in enumerate(docs):
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        # IIF date: MM/DD/YYYY
        raw_date = doc.get("document_date", "")
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            iif_date = dt.strftime("%m/%d/%Y")
        except Exception:
            iif_date = raw_date
        vendor = doc.get("vendor", "")
        doc_id = doc.get("document_id", "")
        memo = f"{vendor} expense"
        gl = doc.get("gl_account", "")

        # TRNS line — the total (Accounts Payable credit)
        lines.append(
            f"TRNS\t{i+1}\tBILL\t{iif_date}\tAccounts Payable\t{vendor}\t"
            f"-{amount}\t{doc_id}\t{memo}\tN\tN\t"
        )
        # SPL line — the expense account debit
        lines.append(
            f"SPL\t{i+1}\tBILL\t{iif_date}\t{gl}\t{taxes['pre_tax']}\t"
            f"{doc_id}\t{memo}\tN\t\t\t"
        )
        # SPL lines for taxes if applicable
        if taxes["gst"] > _ZERO:
            lines.append(
                f"SPL\t{i+1}\tBILL\t{iif_date}\tGST Paid\t{taxes['gst']}\t"
                f"{doc_id}\tGST\tN\t\t\t"
            )
        if taxes["qst"] > _ZERO:
            lines.append(
                f"SPL\t{i+1}\tBILL\t{iif_date}\tQST Paid\t{taxes['qst']}\t"
                f"{doc_id}\tQST\tN\t\t\t"
            )
        if taxes["hst"] > _ZERO:
            lines.append(
                f"SPL\t{i+1}\tBILL\t{iif_date}\tHST Paid\t{taxes['hst']}\t"
                f"{doc_id}\tHST\tN\t\t\t"
            )
        lines.append("ENDTRNS")

    return "\r\n".join(lines).encode("utf-8")


# ---------------------------------------------------------------------------
# PART 5 — Xero CSV Export
# ---------------------------------------------------------------------------

def generate_xero(docs: list[dict[str, Any]]) -> bytes:
    """Generate Xero bank statement import CSV."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Date", "Amount", "Payee", "Description",
        "Reference", "Cheque Number", "Analysed Amount",
    ])
    for doc in docs:
        amount = _dec(doc.get("amount"))
        raw_date = doc.get("document_date", "")
        # Xero date: DD/MM/YYYY
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            xero_date = dt.strftime("%d/%m/%Y")
        except Exception:
            xero_date = raw_date
        writer.writerow([
            sanitize_csv_cell(xero_date),
            str(amount),
            sanitize_csv_cell(doc.get("vendor", "")),
            sanitize_csv_cell(f"{doc.get('vendor', '')} expense"),
            sanitize_csv_cell(doc.get("document_id", "")),
            "",
            str(amount),
        ])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PART 6 — Wave Accounting Export
# ---------------------------------------------------------------------------

def generate_wave(docs: list[dict[str, Any]]) -> bytes:
    """Generate Wave Accounting import CSV."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Transaction Date", "Description", "Debit", "Credit",
        "Account Name", "Tax Name", "Tax Amount",
    ])
    for doc in docs:
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        tc = (doc.get("tax_code", "") or "").strip().upper()
        tax_name = ""
        tax_amount = _ZERO
        if tc in ("T", "GST_QST", "M"):
            tax_name = "GST/QST"
            tax_amount = taxes["gst"] + taxes["qst"]
        elif tc in ("HST", "HST_ATL"):
            tax_name = "HST"
            tax_amount = taxes["hst"]
        writer.writerow([
            sanitize_csv_cell(doc.get("document_date", "")),
            sanitize_csv_cell(f"{doc.get('vendor', '')} expense"),
            str(taxes["pre_tax"]),
            "",
            sanitize_csv_cell(doc.get("gl_account", "")),
            sanitize_csv_cell(tax_name),
            str(tax_amount),
        ])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# PART 7 — Excel Export (multi-sheet)
# ---------------------------------------------------------------------------

def generate_excel(docs: list[dict[str, Any]], client_code: str, period: str) -> bytes:
    """Generate Excel workbook with 4 sheets using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    # OtoCPA blue
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _style_header(ws: Any, headers: list[str]) -> None:
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        ws.freeze_panes = "A2"

    def _auto_width(ws: Any) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # --- Sheet 1: Transactions ---
    ws1 = wb.active
    ws1.title = "Transactions"
    txn_headers = [
        "Date", "Vendor", "Description", "GL Account", "Amount",
        "GST", "QST", "HST", "Tax Code", "Document ID",
    ]
    _style_header(ws1, txn_headers)
    for row_idx, doc in enumerate(docs, 2):
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        ws1.cell(row=row_idx, column=1, value=doc.get("document_date", ""))
        ws1.cell(row=row_idx, column=2, value=doc.get("vendor", ""))
        ws1.cell(row=row_idx, column=3, value=f"{doc.get('vendor', '')} expense")
        ws1.cell(row=row_idx, column=4, value=doc.get("gl_account", ""))
        ws1.cell(row=row_idx, column=5, value=float(amount))
        ws1.cell(row=row_idx, column=6, value=float(taxes["gst"]))
        ws1.cell(row=row_idx, column=7, value=float(taxes["qst"]))
        ws1.cell(row=row_idx, column=8, value=float(taxes["hst"]))
        ws1.cell(row=row_idx, column=9, value=doc.get("tax_code", ""))
        ws1.cell(row=row_idx, column=10, value=doc.get("document_id", ""))
    _auto_width(ws1)

    # --- Sheet 2: GST/QST Summary ---
    ws2 = wb.create_sheet("GST-QST Summary")
    _style_header(ws2, ["Tax Code", "Count", "Total Amount", "GST", "QST", "HST"])
    tax_summary: dict[str, dict[str, Any]] = {}
    for doc in docs:
        tc = (doc.get("tax_code", "") or "").strip().upper() or "NONE"
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        if tc not in tax_summary:
            tax_summary[tc] = {"count": 0, "total": _ZERO, "gst": _ZERO, "qst": _ZERO, "hst": _ZERO}
        tax_summary[tc]["count"] += 1
        tax_summary[tc]["total"] += amount
        tax_summary[tc]["gst"] += taxes["gst"]
        tax_summary[tc]["qst"] += taxes["qst"]
        tax_summary[tc]["hst"] += taxes["hst"]
    for row_idx, (tc, s) in enumerate(sorted(tax_summary.items()), 2):
        ws2.cell(row=row_idx, column=1, value=tc)
        ws2.cell(row=row_idx, column=2, value=s["count"])
        ws2.cell(row=row_idx, column=3, value=float(s["total"]))
        ws2.cell(row=row_idx, column=4, value=float(s["gst"]))
        ws2.cell(row=row_idx, column=5, value=float(s["qst"]))
        ws2.cell(row=row_idx, column=6, value=float(s["hst"]))
    _auto_width(ws2)

    # --- Sheet 3: Trial Balance ---
    ws3 = wb.create_sheet("Trial Balance")
    _style_header(ws3, ["GL Account", "Debit", "Credit", "Net Balance"])
    gl_totals: dict[str, Decimal] = {}
    for doc in docs:
        gl = doc.get("gl_account", "") or "Unassigned"
        amount = _dec(doc.get("amount"))
        taxes = _extract_taxes(amount, doc.get("tax_code", ""))
        gl_totals[gl] = gl_totals.get(gl, _ZERO) + taxes["pre_tax"]
    for row_idx, (gl, total) in enumerate(sorted(gl_totals.items()), 2):
        ws3.cell(row=row_idx, column=1, value=gl)
        ws3.cell(row=row_idx, column=2, value=float(total))
        ws3.cell(row=row_idx, column=3, value=0.0)
        ws3.cell(row=row_idx, column=4, value=float(total))
    _auto_width(ws3)

    # --- Sheet 4: GL Detail ---
    ws4 = wb.create_sheet("GL Detail")
    _style_header(ws4, ["GL Account", "Date", "Vendor", "Amount", "Tax Code", "Document ID"])
    gl_docs: dict[str, list[dict]] = {}
    for doc in docs:
        gl = doc.get("gl_account", "") or "Unassigned"
        gl_docs.setdefault(gl, []).append(doc)
    row_idx = 2
    for gl in sorted(gl_docs.keys()):
        for doc in gl_docs[gl]:
            amount = _dec(doc.get("amount"))
            taxes = _extract_taxes(amount, doc.get("tax_code", ""))
            ws4.cell(row=row_idx, column=1, value=gl)
            ws4.cell(row=row_idx, column=2, value=doc.get("document_date", ""))
            ws4.cell(row=row_idx, column=3, value=doc.get("vendor", ""))
            ws4.cell(row=row_idx, column=4, value=float(taxes["pre_tax"]))
            ws4.cell(row=row_idx, column=5, value=doc.get("tax_code", ""))
            ws4.cell(row=row_idx, column=6, value=doc.get("document_id", ""))
            row_idx += 1
    _auto_width(ws4)

    # Title metadata
    wb.properties.title = f"OtoCPA Export — {client_code} — {period}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PART 9 — Bulk ZIP Export (annual)
# ---------------------------------------------------------------------------

def generate_annual_zip(
    client_code: str,
    year: int,
    db_path: Path = DB_PATH,
) -> bytes:
    """Generate a ZIP with monthly CSVs, annual Excel, and annual summaries."""
    buf = io.BytesIO()
    all_year_docs: list[dict[str, Any]] = []

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for month in range(1, 13):
            period = f"{year:04d}-{month:02d}"
            start, end = _period_dates(period)
            docs = fetch_posted_documents(client_code, start, end, db_path)
            all_year_docs.extend(docs)
            csv_bytes = generate_csv(docs)
            zf.writestr(f"OtoCPA_Export_{client_code}_{period}.csv", csv_bytes)

        # Annual Excel
        excel_bytes = generate_excel(all_year_docs, client_code, str(year))
        zf.writestr(f"OtoCPA_{client_code}_{year}.xlsx", excel_bytes)

        # Annual GST/QST summary CSV
        summary_buf = io.StringIO()
        sw = csv.writer(summary_buf)
        sw.writerow(["Month", "Documents", "Total Amount", "GST", "QST", "HST"])
        for month in range(1, 13):
            period = f"{year:04d}-{month:02d}"
            start, end = _period_dates(period)
            month_docs = [d for d in all_year_docs
                          if start <= (d.get("document_date", "") or "") <= end]
            total_amt = _ZERO
            total_gst = _ZERO
            total_qst = _ZERO
            total_hst = _ZERO
            for doc in month_docs:
                amt = _dec(doc.get("amount"))
                taxes = _extract_taxes(amt, doc.get("tax_code", ""))
                total_amt += amt
                total_gst += taxes["gst"]
                total_qst += taxes["qst"]
                total_hst += taxes["hst"]
            sw.writerow([
                period, len(month_docs),
                str(total_amt), str(total_gst), str(total_qst), str(total_hst),
            ])
        zf.writestr(
            f"OtoCPA_GST_QST_Summary_{client_code}_{year}.csv",
            b"\xef\xbb\xbf" + summary_buf.getvalue().encode("utf-8"),
        )

        # Annual Trial Balance CSV
        tb_buf = io.StringIO()
        tw = csv.writer(tb_buf)
        tw.writerow(["GL Account", "Total Debits", "Total Credits", "Net Balance"])
        gl_totals: dict[str, Decimal] = {}
        for doc in all_year_docs:
            gl = doc.get("gl_account", "") or "Unassigned"
            amt = _dec(doc.get("amount"))
            taxes = _extract_taxes(amt, doc.get("tax_code", ""))
            gl_totals[gl] = gl_totals.get(gl, _ZERO) + taxes["pre_tax"]
        for gl in sorted(gl_totals.keys()):
            tw.writerow([gl, str(gl_totals[gl]), "0.00", str(gl_totals[gl])])
        zf.writestr(
            f"OtoCPA_Trial_Balance_{client_code}_{year}.csv",
            b"\xef\xbb\xbf" + tb_buf.getvalue().encode("utf-8"),
        )

    return buf.getvalue()
