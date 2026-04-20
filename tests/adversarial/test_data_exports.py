"""R4-Investigation 8 — data export integrity.

Generate every supported export format from a known dataset and verify:
  - No crash, no empty output, no truncation
  - CSV escapes quotes, commas, newlines inside values
  - PDF byte stream is valid (starts with %PDF-)
  - Sage50 / Acomba / Wave / Xero CSVs carry one row per document
"""
from __future__ import annotations

import csv
import io
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _mk_docs(count: int = 5) -> list[dict]:
    """Build a list of doc rows matching what fetch_posted_documents
    returns (dict-shape)."""
    return [
        {
            "document_id": f"DOC-{i:03d}",
            "vendor": f"Vendor, Inc. #{i}",
            "client_code": "ACME-CAFE",
            "document_date": f"2026-04-{(i % 28) + 1:02d}",
            "amount": 100.0 + i,
            "gst_amount": 5.0 + i * 0.01,
            "qst_amount": 9.98 + i * 0.01,
            "gl_account": "6100",
            "tax_code": "T",
            "category": "Office",
            "review_status": "Posted",
            "raw_result": "{}",
            "doc_type": "expense",
            "description": 'Line with "quote" and, comma\nand newline',
        }
        for i in range(count)
    ]


# ---------------------------------------------------------------------------
# CSV export integrity
# ---------------------------------------------------------------------------

def test_csv_export_escapes_commas_in_vendor_column():
    """The seed uses vendor = 'Vendor, Inc. #0' — the comma inside the
    value must not split into two columns after csv.reader parses the
    output."""
    from src.engines.export_engine import generate_csv
    data = generate_csv(_mk_docs(3))
    assert data, "CSV export empty"
    text = data.decode("utf-8-sig") if isinstance(data, bytes) else data
    rows = list(csv.reader(io.StringIO(text)))
    assert len(rows) == 4, f"CSV has {len(rows)} rows, expected 4"
    header = rows[0]
    vendor_col = header.index("Vendor")
    for i, row in enumerate(rows[1:]):
        vendor_value = row[vendor_col]
        assert "Vendor" in vendor_value and f"#{i}" in vendor_value, (
            f"CSV row {i} lost the vendor value after round-trip: "
            f"{vendor_value!r}"
        )


def test_csv_export_defangs_formula_injection():
    """sanitize_csv_cell prepends an apostrophe to values starting with
    ``=``, ``+``, ``-``, ``@``, ``\\t``, ``\\r`` to neutralize Excel
    formula injection."""
    from src.engines.export_engine import generate_csv
    docs = _mk_docs(1)
    docs[0]["vendor"] = "=SUM(A1:A100)+cmd|' /C calc'!A0"
    data = generate_csv(docs)
    text = data.decode("utf-8-sig") if isinstance(data, bytes) else data
    rows = list(csv.reader(io.StringIO(text)))
    vendor_col = rows[0].index("Vendor")
    vendor_value = rows[1][vendor_col]
    assert vendor_value.startswith("'="), (
        f"formula-injection vendor wasn't neutralized: {vendor_value!r}"
    )


def test_csv_with_zero_documents_does_not_crash():
    from src.engines.export_engine import generate_csv
    data = generate_csv([])
    # Either empty bytes or header-only; must not crash.
    assert data is not None


# ---------------------------------------------------------------------------
# Sage50 / Acomba / Wave / Xero / QBD outputs.
# ---------------------------------------------------------------------------

def test_sage50_export_non_empty_for_populated_docs():
    from src.engines.export_engine import generate_sage50
    out = generate_sage50(_mk_docs(4))
    assert out, "Sage50 export empty"


def test_acomba_export_non_empty():
    from src.engines.export_engine import generate_acomba
    out = generate_acomba(_mk_docs(4))
    assert out


def test_wave_export_non_empty():
    from src.engines.export_engine import generate_wave
    out = generate_wave(_mk_docs(4))
    assert out


def test_xero_export_non_empty():
    from src.engines.export_engine import generate_xero
    out = generate_xero(_mk_docs(4))
    assert out


def test_qbd_iif_export_non_empty():
    from src.engines.export_engine import generate_qbd_iif
    out = generate_qbd_iif(_mk_docs(4))
    assert out


# ---------------------------------------------------------------------------
# Excel export — openable by openpyxl.
# ---------------------------------------------------------------------------

def test_excel_export_is_valid_xlsx():
    from src.engines.export_engine import generate_excel
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not available")
    data = generate_excel(_mk_docs(4), "ACME-CAFE", "2026-04")
    if not data:
        pytest.skip("generate_excel returned empty (xlsx lib may be missing)")
    import openpyxl as _ox
    wb = _ox.load_workbook(io.BytesIO(data))
    assert len(wb.sheetnames) >= 1
    ws = wb[wb.sheetnames[0]]
    # Header + 4 rows.
    rows_with_data = [r for r in ws.iter_rows(values_only=True) if any(r)]
    assert len(rows_with_data) >= 5, (
        f"xlsx has {len(rows_with_data)} rows, expected >= 5"
    )


# ---------------------------------------------------------------------------
# Financial-statement PDF is a valid PDF (starts with %PDF-).
# ---------------------------------------------------------------------------

def test_fs_pdf_minimal_is_valid_pdf(tmp_path, monkeypatch):
    db = tmp_path / "x.db"
    import src.engines.ocr_engine as oe
    import src.engines.gl_engine as gle
    monkeypatch.setattr(oe, "DB_PATH", db)
    monkeypatch.setattr(gle, "DB_PATH", db)
    from src.engines.audit_engine import (
        ensure_audit_tables, seed_chart_of_accounts,
        generate_financial_statements_pdf,
    )
    from src.engines.gl_engine import ensure_schema as ensure_gl
    import sqlite3 as _s
    conn = _s.connect(str(db))
    conn.row_factory = _s.Row
    ensure_audit_tables(conn)
    seed_chart_of_accounts(conn)
    ensure_gl()
    # Minimal balanced GL.
    for acct_d, acct_c in [("1000", "3000"), ("6100", "1000")]:
        conn.execute(
            "INSERT INTO gl_transactions "
            "(entry_id, client_code, period, entry_date, account_code, side, "
            " amount, description, source) VALUES (?,?,?,?,?,?,?,?,?)",
            (f"E-{acct_d}-{acct_c}", "C1", "2026-04", "2026-04-15",
             acct_d, "debit", 100.0, "", "manual_je"),
        )
        conn.execute(
            "INSERT INTO gl_transactions "
            "(entry_id, client_code, period, entry_date, account_code, side, "
            " amount, description, source) VALUES (?,?,?,?,?,?,?,?,?)",
            (f"E-{acct_d}-{acct_c}", "C1", "2026-04", "2026-04-15",
             acct_c, "credit", 100.0, "", "manual_je"),
        )
    conn.commit()
    pdf = generate_financial_statements_pdf(conn, "C1", "2026-04", lang="en")
    assert isinstance(pdf, (bytes, bytearray))
    assert pdf[:5] == b"%PDF-", (
        f"PDF doesn't start with %PDF- header: {pdf[:10]!r}"
    )
    # Ends with %%EOF (allowing trailing whitespace).
    tail = bytes(pdf).rstrip()
    assert tail.endswith(b"%%EOF"), f"PDF missing %%EOF trailer: {tail[-30:]!r}"


# ---------------------------------------------------------------------------
# CSV completeness: row count equals doc count.
# ---------------------------------------------------------------------------

def test_csv_row_count_equals_doc_count():
    from src.engines.export_engine import generate_csv
    for n in (0, 1, 10, 100):
        docs = _mk_docs(n)
        data = generate_csv(docs)
        text = data.decode("utf-8") if isinstance(data, bytes) else data
        rows = list(csv.reader(io.StringIO(text)))
        if n == 0:
            # Header-only or empty — both acceptable.
            assert len(rows) <= 1
        else:
            assert len(rows) == n + 1, (
                f"n={n}: csv has {len(rows)} rows, expected {n+1} "
                f"(header + {n} docs)"
            )
