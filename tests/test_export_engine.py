"""Tests for src/engines/export_engine.py — all export formats."""
from __future__ import annotations

import csv
import io
import sqlite3
import zipfile
from decimal import Decimal
from pathlib import Path

import pytest

from src.engines.export_engine import (
    _dec,
    _extract_taxes,
    _period_dates,
    fetch_posted_documents,
    generate_acomba,
    generate_csv,
    generate_qbd_iif,
    generate_sage50,
    generate_wave,
    generate_xero,
    generate_annual_zip,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _setup_db(tmp_path: Path) -> Path:
    """Create a minimal DB with documents + posting_jobs tables and seed data."""
    db_path = tmp_path / "test.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE documents (
            document_id TEXT PRIMARY KEY,
            file_name TEXT,
            file_path TEXT,
            client_code TEXT,
            vendor TEXT,
            document_date TEXT,
            amount REAL,
            doc_type TEXT,
            category TEXT,
            gl_account TEXT,
            tax_code TEXT,
            review_status TEXT,
            confidence REAL,
            raw_result TEXT,
            created_at TEXT,
            updated_at TEXT,
            assigned_to TEXT,
            manual_hold_reason TEXT,
            manual_hold_by TEXT,
            manual_hold_at TEXT,
            fraud_flags TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE posting_jobs (
            posting_id TEXT PRIMARY KEY,
            document_id TEXT,
            target_system TEXT,
            posting_status TEXT,
            approval_state TEXT,
            reviewer TEXT,
            external_id TEXT,
            payload_json TEXT,
            error_text TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    # Insert test documents
    docs = [
        ("DOC001", "CLIENT1", "Vendor A", "2026-01-15", 114.98, "5000", "T", "posted", "EXT1"),
        ("DOC002", "CLIENT1", "Vendor B", "2026-01-20", 50.00, "5100", "E", "posted", "EXT2"),
        ("DOC003", "CLIENT1", "Vendor C", "2026-02-10", 200.00, "5200", "HST", "posted", "EXT3"),
        ("DOC004", "CLIENT1", "Vendor D", "2026-01-25", 75.00, "5300", "M", "posted", "EXT4"),
        ("DOC005", "CLIENT1", "Vendor E", "2026-01-05", 30.00, "5400", "Z", "posted", "EXT5"),
    ]
    for doc_id, client, vendor, date, amount, gl, tc, _, _ in docs:
        conn.execute(
            "INSERT INTO documents (document_id, client_code, vendor, document_date, amount, gl_account, tax_code, review_status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 'Ready')",
            (doc_id, client, vendor, date, amount, gl, tc),
        )
    for doc_id, _, _, _, _, _, _, ps, ext in docs:
        conn.execute(
            "INSERT INTO posting_jobs (posting_id, document_id, posting_status, external_id, created_at) "
            "VALUES (?, ?, ?, ?, datetime('now'))",
            (f"PJ-{doc_id}", doc_id, ps, ext),
        )
    conn.commit()
    conn.close()
    return db_path


SAMPLE_DOCS = [
    {
        "document_id": "D1",
        "vendor": "Fournisseur Québec",
        "document_date": "2026-01-15",
        "amount": "114.98",
        "gl_account": "5000",
        "tax_code": "T",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "inv1.pdf",
        "client_code": "CLI1",
    },
    {
        "document_id": "D2",
        "vendor": "Vendor B",
        "document_date": "2026-01-20",
        "amount": "50.00",
        "gl_account": "5100",
        "tax_code": "E",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "inv2.pdf",
        "client_code": "CLI1",
    },
]


# ---------------------------------------------------------------------------
# Unit tests — helpers
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_dec_none(self):
        assert _dec(None) == Decimal("0")

    def test_dec_empty(self):
        assert _dec("") == Decimal("0")

    def test_dec_valid(self):
        assert _dec("114.98") == Decimal("114.98")

    def test_dec_invalid(self):
        assert _dec("abc") == Decimal("0")

    def test_period_dates_january(self):
        start, end = _period_dates("2026-01")
        assert start == "2026-01-01"
        assert end == "2026-01-31"

    def test_period_dates_february_non_leap(self):
        start, end = _period_dates("2025-02")
        assert start == "2025-02-01"
        assert end == "2025-02-28"

    def test_period_dates_february_leap(self):
        start, end = _period_dates("2028-02")
        assert start == "2028-02-01"
        assert end == "2028-02-29"

    def test_extract_taxes_taxable(self):
        taxes = _extract_taxes(Decimal("114.98"), "T")
        assert taxes["gst"] > Decimal("0")
        assert taxes["qst"] > Decimal("0")
        assert taxes["hst"] == Decimal("0")
        # pre_tax + gst + qst should approximate total
        total = taxes["pre_tax"] + taxes["gst"] + taxes["qst"]
        assert abs(total - Decimal("114.98")) <= Decimal("0.02")

    def test_extract_taxes_exempt(self):
        taxes = _extract_taxes(Decimal("50.00"), "E")
        assert taxes["gst"] == Decimal("0")
        assert taxes["qst"] == Decimal("0")
        assert taxes["pre_tax"] == Decimal("50.00")

    def test_extract_taxes_hst(self):
        taxes = _extract_taxes(Decimal("113.00"), "HST")
        assert taxes["hst"] > Decimal("0")
        assert taxes["gst"] == Decimal("0")
        assert taxes["qst"] == Decimal("0")

    def test_extract_taxes_zero_rated(self):
        taxes = _extract_taxes(Decimal("100.00"), "Z")
        assert taxes["pre_tax"] == Decimal("100.00")
        assert taxes["gst"] == Decimal("0")


# ---------------------------------------------------------------------------
# CSV format tests
# ---------------------------------------------------------------------------

class TestCSVExport:
    def test_csv_has_bom(self):
        data = generate_csv(SAMPLE_DOCS)
        assert data[:3] == b"\xef\xbb\xbf"

    def test_csv_header_row(self):
        data = generate_csv(SAMPLE_DOCS)
        text = data[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert header == [
            "Date", "Vendor", "Description", "GL Account", "Amount",
            "GST", "QST", "HST", "Tax Code", "Document ID",
        ]

    def test_csv_row_count(self):
        data = generate_csv(SAMPLE_DOCS)
        text = data[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 3  # header + 2 data rows

    def test_csv_french_characters(self):
        data = generate_csv(SAMPLE_DOCS)
        text = data.decode("utf-8-sig")
        assert "Fournisseur Québec" in text

    def test_csv_empty_docs(self):
        data = generate_csv([])
        text = data[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 1  # header only


# ---------------------------------------------------------------------------
# Sage 50 format tests
# ---------------------------------------------------------------------------

class TestSage50Export:
    def test_sage50_header(self):
        data = generate_sage50(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert header == [
            "Date", "Reference", "Description", "Account Number",
            "Debit", "Credit", "Tax Code",
        ]

    def test_sage50_date_format(self):
        data = generate_sage50(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        next(reader)  # skip header
        row = next(reader)
        # Date should be MM/DD/YYYY
        assert row[0] == "01/15/2026"

    def test_sage50_tax_code_mapping(self):
        data = generate_sage50(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        next(reader)
        row1 = next(reader)  # T -> GP
        row2 = next(reader)  # E -> E
        assert row1[6] == "GP"
        assert row2[6] == "E"


# ---------------------------------------------------------------------------
# Acomba format tests
# ---------------------------------------------------------------------------

class TestAcombaExport:
    def test_acomba_tab_delimited(self):
        data = generate_acomba(SAMPLE_DOCS)
        text = data.decode("utf-8")
        lines = text.strip().split("\n")
        # Tab-separated
        assert "\t" in lines[0]

    def test_acomba_french_headers(self):
        data = generate_acomba(SAMPLE_DOCS)
        text = data.decode("utf-8")
        first_line = text.split("\n")[0]
        assert "No_Pièce" in first_line
        assert "Débit" in first_line
        assert "Crédit" in first_line
        assert "TPS" in first_line
        assert "TVQ" in first_line

    def test_acomba_date_format(self):
        data = generate_acomba(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        next(reader)
        row = next(reader)
        assert row[1] == "20260115"  # YYYYMMDD


# ---------------------------------------------------------------------------
# QuickBooks Desktop IIF tests
# ---------------------------------------------------------------------------

class TestQBDExport:
    def test_iif_header_lines(self):
        data = generate_qbd_iif(SAMPLE_DOCS)
        text = data.decode("utf-8")
        lines = text.split("\r\n")
        assert lines[0].startswith("!TRNS")
        assert lines[1].startswith("!SPL")
        assert lines[2] == "!ENDTRNS"

    def test_iif_trns_and_spl(self):
        data = generate_qbd_iif(SAMPLE_DOCS)
        text = data.decode("utf-8")
        assert "TRNS\t" in text
        assert "SPL\t" in text
        assert "ENDTRNS" in text

    def test_iif_has_tax_lines(self):
        # First doc is taxable (T), should have GST and QST SPL lines
        data = generate_qbd_iif(SAMPLE_DOCS[:1])
        text = data.decode("utf-8")
        assert "GST Paid" in text
        assert "QST Paid" in text

    def test_iif_exempt_no_tax_lines(self):
        # Second doc is exempt (E), no tax lines
        data = generate_qbd_iif(SAMPLE_DOCS[1:2])
        text = data.decode("utf-8")
        assert "GST Paid" not in text
        assert "QST Paid" not in text


# ---------------------------------------------------------------------------
# Xero format tests
# ---------------------------------------------------------------------------

class TestXeroExport:
    def test_xero_header(self):
        data = generate_xero(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert "Date" in header
        assert "Payee" in header
        assert "Analysed Amount" in header

    def test_xero_date_format(self):
        data = generate_xero(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        next(reader)
        row = next(reader)
        assert row[0] == "15/01/2026"  # DD/MM/YYYY


# ---------------------------------------------------------------------------
# Wave format tests
# ---------------------------------------------------------------------------

class TestWaveExport:
    def test_wave_header(self):
        data = generate_wave(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert header == [
            "Transaction Date", "Description", "Debit", "Credit",
            "Account Name", "Tax Name", "Tax Amount",
        ]

    def test_wave_tax_name(self):
        data = generate_wave(SAMPLE_DOCS)
        text = data.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        next(reader)
        row1 = next(reader)  # T -> GST/QST
        row2 = next(reader)  # E -> empty
        assert row1[5] == "GST/QST"
        assert row2[5] == ""


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------

class TestExcelExport:
    def test_excel_returns_bytes(self):
        try:
            from src.engines.export_engine import generate_excel
        except ImportError:
            pytest.skip("openpyxl not installed")
        data = generate_excel(SAMPLE_DOCS, "CLI1", "2026-01")
        assert isinstance(data, bytes)
        assert len(data) > 100

    def test_excel_has_four_sheets(self):
        try:
            from openpyxl import load_workbook
            from src.engines.export_engine import generate_excel
        except ImportError:
            pytest.skip("openpyxl not installed")
        data = generate_excel(SAMPLE_DOCS, "CLI1", "2026-01")
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.sheetnames) == 4
        assert wb.sheetnames[0] == "Transactions"
        assert wb.sheetnames[1] == "GST-QST Summary"
        assert wb.sheetnames[2] == "Trial Balance"
        assert wb.sheetnames[3] == "GL Detail"

    def test_excel_transaction_count(self):
        try:
            from openpyxl import load_workbook
            from src.engines.export_engine import generate_excel
        except ImportError:
            pytest.skip("openpyxl not installed")
        data = generate_excel(SAMPLE_DOCS, "CLI1", "2026-01")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Transactions"]
        # header + 2 data rows
        assert ws.max_row == 3

    def test_excel_empty_docs(self):
        try:
            from src.engines.export_engine import generate_excel
        except ImportError:
            pytest.skip("openpyxl not installed")
        data = generate_excel([], "CLI1", "2026-01")
        assert isinstance(data, bytes)


# ---------------------------------------------------------------------------
# DB fetch tests
# ---------------------------------------------------------------------------

class TestFetchPostedDocuments:
    def test_fetch_returns_posted_only(self, tmp_path):
        db_path = _setup_db(tmp_path)
        docs = fetch_posted_documents("CLIENT1", "2026-01-01", "2026-01-31", db_path)
        # Should return DOC001, DOC002, DOC004, DOC005 (all January posted)
        doc_ids = {d["document_id"] for d in docs}
        assert "DOC001" in doc_ids
        assert "DOC002" in doc_ids
        assert "DOC003" not in doc_ids  # February
        assert "DOC004" in doc_ids
        assert "DOC005" in doc_ids

    def test_fetch_case_insensitive(self, tmp_path):
        db_path = _setup_db(tmp_path)
        docs = fetch_posted_documents("client1", "2026-01-01", "2026-01-31", db_path)
        assert len(docs) == 4

    def test_fetch_missing_db(self, tmp_path):
        docs = fetch_posted_documents("X", "2026-01-01", "2026-01-31", tmp_path / "no.db")
        assert docs == []


# ---------------------------------------------------------------------------
# Annual ZIP tests
# ---------------------------------------------------------------------------

class TestAnnualZip:
    def test_zip_structure(self, tmp_path):
        db_path = _setup_db(tmp_path)
        data = generate_annual_zip("CLIENT1", 2026, db_path)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
            # 12 monthly CSVs + 1 Excel + 1 GST summary + 1 trial balance
            assert len(names) == 15
            # Check specific files exist
            assert "OtoCPA_Export_CLIENT1_2026-01.csv" in names
            assert "OtoCPA_Export_CLIENT1_2026-12.csv" in names
            assert "OtoCPA_CLIENT1_2026.xlsx" in names
            assert "OtoCPA_GST_QST_Summary_CLIENT1_2026.csv" in names
            assert "OtoCPA_Trial_Balance_CLIENT1_2026.csv" in names

    def test_zip_is_valid(self, tmp_path):
        db_path = _setup_db(tmp_path)
        data = generate_annual_zip("CLIENT1", 2026, db_path)
        assert zipfile.is_zipfile(io.BytesIO(data))

    def test_zip_empty_db(self, tmp_path):
        # Works even with no data
        db_path = _setup_db(tmp_path)
        data = generate_annual_zip("NONEXISTENT", 2026, db_path)
        assert zipfile.is_zipfile(io.BytesIO(data))
