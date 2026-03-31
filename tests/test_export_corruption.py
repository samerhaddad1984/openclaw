"""
tests/test_export_corruption.py — Export corruption detection suite.

Generates full books, exports to every supported format, and verifies:
- Totals preserved (pre_tax + gst + qst == amount within rounding)
- Accents / French text survive round-trip
- Tax codes map correctly per format
- Locked periods do not export edited values
- Signs never flip
- Encoding never corrupts
- Customer/vendor IDs never mutate
- Rounding stays ROUND_HALF_UP to 0.01

Fails if any export changes sign, encoding, tax mapping, customer/vendor IDs,
or rounding.
"""
from __future__ import annotations

import csv
import io
import sqlite3
import zipfile
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import pytest

from src.engines.export_engine import (
    CENT,
    GST_RATE,
    QST_RATE,
    _dec,
    _extract_taxes,
    _round,
    generate_acomba,
    generate_csv,
    generate_qbd_iif,
    generate_sage50,
    generate_wave,
    generate_xero,
    generate_annual_zip,
    fetch_posted_documents,
)

_ZERO = Decimal("0")


# ---------------------------------------------------------------------------
# Fixtures — full book with diverse tax codes, French vendors, edge amounts
# ---------------------------------------------------------------------------

FULL_BOOKS = [
    {
        "document_id": "INV-001",
        "vendor": "Fournisseur Québec Ltée",
        "document_date": "2026-01-10",
        "amount": "114.98",
        "gl_account": "5000",
        "tax_code": "T",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "facture_001.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-002",
        "vendor": "Café Éléphant Résolu",
        "document_date": "2026-01-15",
        "amount": "229.95",
        "gl_account": "5100",
        "tax_code": "GST_QST",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "facture_002.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-003",
        "vendor": "Hydro-Québec",
        "document_date": "2026-01-20",
        "amount": "350.00",
        "gl_account": "5200",
        "tax_code": "E",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "hydro.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-004",
        "vendor": "Ministère du Revenu",
        "document_date": "2026-01-25",
        "amount": "0.01",
        "gl_account": "5300",
        "tax_code": "Z",
        "category": "expense",
        "doc_type": "receipt",
        "file_name": "penny.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-005",
        "vendor": "Ontario Supplies Inc.",
        "document_date": "2026-02-05",
        "amount": "113.00",
        "gl_account": "5400",
        "tax_code": "HST",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "ont_supply.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-006",
        "vendor": "Atlantic Côtière Ltd",
        "document_date": "2026-02-10",
        "amount": "115.00",
        "gl_account": "5500",
        "tax_code": "HST_ATL",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "atlantic.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        "document_id": "INV-007",
        "vendor": "Réparations Générales Noël",
        "document_date": "2026-02-15",
        "amount": "575.50",
        "gl_account": "5000",
        "tax_code": "M",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "reparations.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        # Zero-amount edge case
        "document_id": "INV-008",
        "vendor": "Crédit Ajustement",
        "document_date": "2026-03-01",
        "amount": "0.00",
        "gl_account": "5600",
        "tax_code": "T",
        "category": "adjustment",
        "doc_type": "credit_note",
        "file_name": "zero.pdf",
        "client_code": "DÉMOCLI",
    },
    {
        # Large amount for rounding stress
        "document_id": "INV-009",
        "vendor": "Grossiste Méga Approvisionnement",
        "document_date": "2026-03-15",
        "amount": "999999.99",
        "gl_account": "5700",
        "tax_code": "T",
        "category": "expense",
        "doc_type": "invoice",
        "file_name": "mega.pdf",
        "client_code": "DÉMOCLI",
    },
]

# All French-accented vendor names from the test data
FRENCH_VENDORS = [
    "Fournisseur Québec Ltée",
    "Café Éléphant Résolu",
    "Hydro-Québec",
    "Ministère du Revenu",
    "Atlantic Côtière Ltd",
    "Réparations Générales Noël",
    "Crédit Ajustement",
    "Grossiste Méga Approvisionnement",
]


def _parse_csv_bytes(data: bytes, bom: bool = False) -> list[list[str]]:
    """Parse CSV bytes into rows, stripping BOM if present."""
    if bom and data[:3] == b"\xef\xbb\xbf":
        data = data[3:]
    text = data.decode("utf-8")
    return list(csv.reader(io.StringIO(text)))


def _parse_tsv_bytes(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8")
    return list(csv.reader(io.StringIO(text), delimiter="\t"))


def _setup_locked_db(tmp_path: Path) -> Path:
    """Create a DB with a locked period (Jan 2026) and an unlocked period (Feb 2026)."""
    db_path = tmp_path / "locked.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE documents (
            document_id TEXT PRIMARY KEY,
            file_name TEXT, file_path TEXT, client_code TEXT,
            vendor TEXT, document_date TEXT, amount REAL,
            doc_type TEXT, category TEXT, gl_account TEXT,
            tax_code TEXT, review_status TEXT, confidence REAL,
            raw_result TEXT, created_at TEXT, updated_at TEXT,
            assigned_to TEXT, manual_hold_reason TEXT,
            manual_hold_by TEXT, manual_hold_at TEXT, fraud_flags TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE posting_jobs (
            posting_id TEXT PRIMARY KEY,
            document_id TEXT, target_system TEXT, posting_status TEXT,
            approval_state TEXT, reviewer TEXT, external_id TEXT,
            payload_json TEXT, error_text TEXT,
            created_at TEXT, updated_at TEXT
        )
    """)
    # Locked-period doc (posted in Jan)
    conn.execute(
        "INSERT INTO documents VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ("LOCK-001", "f.pdf", "/f.pdf", "CLI1", "Vendeur Verrouillé",
         "2026-01-15", 114.98, "invoice", "expense", "5000", "T",
         "Ready", 0.99, None, "2026-01-15T10:00:00", "2026-01-15T10:00:00",
         None, None, None, None, None),
    )
    conn.execute(
        "INSERT INTO posting_jobs VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        ("PJ-LOCK-001", "LOCK-001", "qbo", "posted", "approved", "admin",
         "EXT-LOCK-001", None, None, "2026-01-15T10:00:00", "2026-01-15T10:00:00"),
    )
    # Unlocked-period doc (posted in Feb)
    conn.execute(
        "INSERT INTO documents VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ("UNLOCK-001", "g.pdf", "/g.pdf", "CLI1", "Vendeur Ouvert",
         "2026-02-10", 200.00, "invoice", "expense", "5100", "E",
         "Ready", 0.95, None, "2026-02-10T10:00:00", "2026-02-10T10:00:00",
         None, None, None, None, None),
    )
    conn.execute(
        "INSERT INTO posting_jobs VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        ("PJ-UNLOCK-001", "UNLOCK-001", "qbo", "posted", "approved", "admin",
         "EXT-UNLOCK-001", None, None, "2026-02-10T10:00:00", "2026-02-10T10:00:00"),
    )
    conn.commit()
    conn.close()
    return db_path


# ===================================================================
# 1. TOTALS PRESERVED — every format must preserve the original amount
# ===================================================================

class TestTotalsPreserved:
    """Export must never lose or gain cents. pre_tax + taxes == original amount (within rounding)."""

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_csv_total_preserved(self, doc):
        rows = _parse_csv_bytes(generate_csv([doc]), bom=True)
        data_row = rows[1]  # first data row
        amount = Decimal(data_row[4])
        gst = Decimal(data_row[5])
        qst = Decimal(data_row[6])
        hst = Decimal(data_row[7])
        original = _dec(doc["amount"])
        assert amount == original, f"CSV amount {amount} != original {original}"
        # Reconstruct: the CSV stores full amount in Amount column,
        # but pre_tax + gst + qst + hst should be close to amount
        taxes = _extract_taxes(original, doc["tax_code"])
        reconstructed = taxes["pre_tax"] + gst + qst + hst
        assert abs(reconstructed - original) <= Decimal("0.02"), (
            f"CSV tax reconstruction drift: {reconstructed} vs {original}"
        )

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_sage50_debit_matches_pretax(self, doc):
        rows = _parse_csv_bytes(generate_sage50([doc]))
        data_row = rows[1]
        debit = Decimal(data_row[4])
        original = _dec(doc["amount"])
        taxes = _extract_taxes(original, doc["tax_code"])
        assert debit == taxes["pre_tax"], f"Sage50 debit {debit} != pre_tax {taxes['pre_tax']}"

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_acomba_debit_matches_pretax(self, doc):
        rows = _parse_tsv_bytes(generate_acomba([doc]))
        data_row = rows[1]
        debit = Decimal(data_row[4])
        original = _dec(doc["amount"])
        taxes = _extract_taxes(original, doc["tax_code"])
        assert debit == taxes["pre_tax"], f"Acomba debit {debit} != pre_tax {taxes['pre_tax']}"

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_xero_amount_preserved(self, doc):
        rows = _parse_csv_bytes(generate_xero([doc]))
        data_row = rows[1]
        exported_amount = Decimal(data_row[1])
        original = _dec(doc["amount"])
        assert exported_amount == original, f"Xero amount {exported_amount} != {original}"
        # Analysed Amount should also match
        analysed = Decimal(data_row[6])
        assert analysed == original, f"Xero analysed {analysed} != {original}"

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_wave_debit_plus_tax_equals_amount(self, doc):
        rows = _parse_csv_bytes(generate_wave([doc]))
        data_row = rows[1]
        debit = Decimal(data_row[2])
        tax_amount = Decimal(data_row[6])
        original = _dec(doc["amount"])
        reconstructed = debit + tax_amount
        assert abs(reconstructed - original) <= Decimal("0.02"), (
            f"Wave total drift: {debit} + {tax_amount} = {reconstructed} vs {original}"
        )

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_iif_trns_amount_matches(self, doc):
        """IIF TRNS line must carry -amount (AP credit), SPL lines must sum to +amount."""
        data = generate_qbd_iif([doc])
        text = data.decode("utf-8")
        lines = text.split("\r\n")
        original = _dec(doc["amount"])

        trns_amount = None
        spl_sum = _ZERO
        for line in lines:
            parts = line.split("\t")
            if parts[0] == "TRNS":
                trns_amount = Decimal(parts[6])
            elif parts[0] == "SPL":
                spl_sum += Decimal(parts[5])

        assert trns_amount is not None, "No TRNS line found"
        assert trns_amount == -original, f"IIF TRNS amount {trns_amount} != -{original}"
        assert abs(spl_sum - original) <= Decimal("0.02"), (
            f"IIF SPL sum {spl_sum} != {original}"
        )


# ===================================================================
# 2. ACCENTS / FRENCH TEXT SURVIVES ROUND-TRIP
# ===================================================================

class TestAccentsPreserved:
    """French characters (é, è, ê, ë, ô, ç, ù, à, î, ï, ü, â, û, Noël) must survive."""

    def test_csv_french_roundtrip(self):
        data = generate_csv(FULL_BOOKS)
        text = data.decode("utf-8-sig")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"CSV lost French vendor: {vendor}"

    def test_sage50_french_roundtrip(self):
        data = generate_sage50(FULL_BOOKS)
        text = data.decode("utf-8")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"Sage50 lost French vendor: {vendor}"

    def test_acomba_french_roundtrip(self):
        data = generate_acomba(FULL_BOOKS)
        text = data.decode("utf-8")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"Acomba lost French vendor: {vendor}"

    def test_iif_french_roundtrip(self):
        data = generate_qbd_iif(FULL_BOOKS)
        text = data.decode("utf-8")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"IIF lost French vendor: {vendor}"

    def test_xero_french_roundtrip(self):
        data = generate_xero(FULL_BOOKS)
        text = data.decode("utf-8")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"Xero lost French vendor: {vendor}"

    def test_wave_french_roundtrip(self):
        data = generate_wave(FULL_BOOKS)
        text = data.decode("utf-8")
        for vendor in FRENCH_VENDORS:
            assert vendor in text, f"Wave lost French vendor: {vendor}"

    def test_excel_french_roundtrip(self):
        try:
            from openpyxl import load_workbook
            from src.engines.export_engine import generate_excel
        except ImportError:
            pytest.skip("openpyxl not installed")
        data = generate_excel(FULL_BOOKS, "DÉMOCLI", "2026-01")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Transactions"]
        exported_vendors = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1]:
                exported_vendors.add(row[1])
        for vendor in FRENCH_VENDORS:
            assert vendor in exported_vendors, f"Excel lost French vendor: {vendor}"

    def test_csv_bom_present_for_excel_compat(self):
        """UTF-8 BOM is required for French characters in Excel."""
        data = generate_csv(FULL_BOOKS)
        assert data[:3] == b"\xef\xbb\xbf", "CSV missing BOM — Excel will mangle accents"

    def test_acomba_french_headers_intact(self):
        data = generate_acomba(FULL_BOOKS)
        text = data.decode("utf-8")
        for header in ("No_Pièce", "Débit", "Crédit", "TPS", "TVQ"):
            assert header in text, f"Acomba lost French header: {header}"


# ===================================================================
# 3. TAX CODES MAP CORRECTLY PER FORMAT
# ===================================================================

class TestTaxCodeMapping:
    """Each format must translate internal tax codes to format-specific codes."""

    SAGE50_EXPECTED = {
        "T": "GP", "GST_QST": "GP", "E": "E", "Z": "Z",
        "HST": "H", "HST_ATL": "H", "M": "GP",
    }

    @pytest.mark.parametrize("tc,expected", list(SAGE50_EXPECTED.items()))
    def test_sage50_tax_code(self, tc, expected):
        doc = {**FULL_BOOKS[0], "tax_code": tc, "document_id": f"TC-{tc}"}
        rows = _parse_csv_bytes(generate_sage50([doc]))
        assert rows[1][6] == expected, f"Sage50 mapped {tc} to {rows[1][6]}, expected {expected}"

    def test_wave_gst_qst_label(self):
        for tc in ("T", "GST_QST", "M"):
            doc = {**FULL_BOOKS[0], "tax_code": tc}
            rows = _parse_csv_bytes(generate_wave([doc]))
            assert rows[1][5] == "GST/QST", f"Wave: tax_code={tc} should map to GST/QST"

    def test_wave_hst_label(self):
        for tc in ("HST", "HST_ATL"):
            doc = {**FULL_BOOKS[0], "tax_code": tc}
            rows = _parse_csv_bytes(generate_wave([doc]))
            assert rows[1][5] == "HST", f"Wave: tax_code={tc} should map to HST"

    def test_wave_exempt_no_tax(self):
        for tc in ("E", "Z"):
            doc = {**FULL_BOOKS[0], "tax_code": tc}
            rows = _parse_csv_bytes(generate_wave([doc]))
            assert rows[1][5] == "", f"Wave: tax_code={tc} should have empty tax name"
            assert Decimal(rows[1][6]) == _ZERO, f"Wave: exempt should have zero tax amount"

    def test_iif_gst_qst_spl_lines(self):
        doc = {**FULL_BOOKS[0], "tax_code": "T", "amount": "114.98"}
        text = generate_qbd_iif([doc]).decode("utf-8")
        assert "GST Paid" in text
        assert "QST Paid" in text
        assert "HST Paid" not in text

    def test_iif_hst_spl_line(self):
        doc = {**FULL_BOOKS[0], "tax_code": "HST", "amount": "113.00"}
        text = generate_qbd_iif([doc]).decode("utf-8")
        assert "HST Paid" in text
        assert "GST Paid" not in text
        assert "QST Paid" not in text

    def test_iif_exempt_no_tax_spl(self):
        doc = {**FULL_BOOKS[0], "tax_code": "E", "amount": "100.00"}
        text = generate_qbd_iif([doc]).decode("utf-8")
        assert "GST Paid" not in text
        assert "QST Paid" not in text
        assert "HST Paid" not in text

    def test_csv_tax_code_passthrough(self):
        """CSV should preserve the original tax code string unmodified."""
        for doc in FULL_BOOKS:
            rows = _parse_csv_bytes(generate_csv([doc]), bom=True)
            assert rows[1][8] == doc["tax_code"], (
                f"CSV mutated tax_code: {rows[1][8]} != {doc['tax_code']}"
            )


# ===================================================================
# 4. LOCKED PERIODS DO NOT EXPORT EDITED VALUES
# ===================================================================

class TestLockedPeriods:
    """
    Locked-period documents must export with their original posted values.
    Simulate by confirming fetch_posted_documents returns data unchanged.
    """

    def test_locked_period_amount_unchanged(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        docs = fetch_posted_documents("CLI1", "2026-01-01", "2026-01-31", db_path)
        assert len(docs) == 1
        assert docs[0]["document_id"] == "LOCK-001"
        assert float(docs[0]["amount"]) == 114.98

    def test_locked_period_exported_amount_matches_db(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        docs = fetch_posted_documents("CLI1", "2026-01-01", "2026-01-31", db_path)
        csv_rows = _parse_csv_bytes(generate_csv(docs), bom=True)
        exported_amount = Decimal(csv_rows[1][4])
        assert exported_amount == Decimal("114.98")

    def test_locked_period_vendor_unchanged(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        docs = fetch_posted_documents("CLI1", "2026-01-01", "2026-01-31", db_path)
        assert docs[0]["vendor"] == "Vendeur Verrouillé"

    def test_unlocked_period_separate(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        jan = fetch_posted_documents("CLI1", "2026-01-01", "2026-01-31", db_path)
        feb = fetch_posted_documents("CLI1", "2026-02-01", "2026-02-28", db_path)
        assert len(jan) == 1
        assert len(feb) == 1
        assert jan[0]["document_id"] == "LOCK-001"
        assert feb[0]["document_id"] == "UNLOCK-001"

    def test_locked_data_not_mutated_by_export(self, tmp_path):
        """Export functions must NOT modify the input dicts."""
        db_path = _setup_locked_db(tmp_path)
        docs = fetch_posted_documents("CLI1", "2026-01-01", "2026-01-31", db_path)
        original = {k: v for k, v in docs[0].items()}
        generate_csv(docs)
        generate_sage50(docs)
        generate_acomba(docs)
        generate_qbd_iif(docs)
        generate_xero(docs)
        generate_wave(docs)
        for key in original:
            assert docs[0][key] == original[key], (
                f"Export mutated input dict key={key}: {docs[0][key]} != {original[key]}"
            )


# ===================================================================
# 5. SIGN NEVER FLIPS
# ===================================================================

class TestSignPreservation:
    """Positive amounts must stay positive in every format. IIF AP line must be negative."""

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_csv_amount_sign(self, doc):
        rows = _parse_csv_bytes(generate_csv([doc]), bom=True)
        amount = Decimal(rows[1][4])
        original = _dec(doc["amount"])
        assert amount >= _ZERO if original >= _ZERO else amount < _ZERO

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_sage50_debit_non_negative(self, doc):
        rows = _parse_csv_bytes(generate_sage50([doc]))
        debit = Decimal(rows[1][4])
        assert debit >= _ZERO, f"Sage50 debit went negative: {debit}"

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_xero_amount_sign(self, doc):
        rows = _parse_csv_bytes(generate_xero([doc]))
        amount = Decimal(rows[1][1])
        original = _dec(doc["amount"])
        assert amount >= _ZERO if original >= _ZERO else amount < _ZERO

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_iif_trns_is_negative_spl_is_positive(self, doc):
        """IIF: TRNS (AP) must be -amount, SPL (expense) must be positive."""
        original = _dec(doc["amount"])
        if original == _ZERO:
            return  # Skip zero amounts
        text = generate_qbd_iif([doc]).decode("utf-8")
        for line in text.split("\r\n"):
            parts = line.split("\t")
            if parts[0] == "TRNS":
                assert Decimal(parts[6]) < _ZERO, "IIF TRNS should be negative (AP credit)"
            elif parts[0] == "SPL":
                assert Decimal(parts[5]) >= _ZERO, "IIF SPL should be positive (expense debit)"

    @pytest.mark.parametrize("doc", FULL_BOOKS, ids=[d["document_id"] for d in FULL_BOOKS])
    def test_wave_debit_non_negative(self, doc):
        rows = _parse_csv_bytes(generate_wave([doc]))
        debit = Decimal(rows[1][2])
        assert debit >= _ZERO, f"Wave debit went negative: {debit}"


# ===================================================================
# 6. ENCODING NEVER CORRUPTS
# ===================================================================

class TestEncodingIntegrity:
    """All exports must decode cleanly with the expected encoding."""

    def test_csv_utf8_bom_decode(self):
        data = generate_csv(FULL_BOOKS)
        text = data.decode("utf-8-sig")  # Must not raise
        assert "Québec" in text

    def test_sage50_utf8_decode(self):
        data = generate_sage50(FULL_BOOKS)
        text = data.decode("utf-8")  # Must not raise
        assert "Québec" in text

    def test_acomba_utf8_decode(self):
        data = generate_acomba(FULL_BOOKS)
        text = data.decode("utf-8")
        assert "Québec" in text

    def test_iif_utf8_decode(self):
        data = generate_qbd_iif(FULL_BOOKS)
        text = data.decode("utf-8")
        assert "Québec" in text

    def test_xero_utf8_decode(self):
        data = generate_xero(FULL_BOOKS)
        text = data.decode("utf-8")
        assert "Québec" in text

    def test_wave_utf8_decode(self):
        data = generate_wave(FULL_BOOKS)
        text = data.decode("utf-8")
        assert "Québec" in text

    def test_no_latin1_mojibake(self):
        """Ensure no double-encoding (UTF-8 bytes re-encoded as Latin-1)."""
        data = generate_csv(FULL_BOOKS)
        text = data.decode("utf-8-sig")
        # Mojibake signature: Ã© instead of é
        assert "Ã©" not in text, "Detected double-encoding mojibake"
        assert "Ã‰" not in text, "Detected double-encoding mojibake"

    def test_csv_no_null_bytes(self):
        data = generate_csv(FULL_BOOKS)
        assert b"\x00" not in data, "CSV contains null bytes"

    def test_iif_no_null_bytes(self):
        data = generate_qbd_iif(FULL_BOOKS)
        assert b"\x00" not in data, "IIF contains null bytes"


# ===================================================================
# 7. CUSTOMER/VENDOR IDS NEVER MUTATE
# ===================================================================

class TestVendorIdPreservation:
    """document_id and vendor must survive export unchanged."""

    def test_csv_document_ids(self):
        rows = _parse_csv_bytes(generate_csv(FULL_BOOKS), bom=True)
        exported_ids = {r[9] for r in rows[1:]}
        expected_ids = {d["document_id"] for d in FULL_BOOKS}
        assert exported_ids == expected_ids

    def test_sage50_document_ids(self):
        rows = _parse_csv_bytes(generate_sage50(FULL_BOOKS))
        exported_ids = {r[1] for r in rows[1:]}
        expected_ids = {d["document_id"] for d in FULL_BOOKS}
        assert exported_ids == expected_ids

    def test_acomba_document_ids(self):
        rows = _parse_tsv_bytes(generate_acomba(FULL_BOOKS))
        exported_ids = {r[0] for r in rows[1:]}
        expected_ids = {d["document_id"] for d in FULL_BOOKS}
        assert exported_ids == expected_ids

    def test_xero_document_ids(self):
        rows = _parse_csv_bytes(generate_xero(FULL_BOOKS))
        exported_ids = {r[4] for r in rows[1:]}
        expected_ids = {d["document_id"] for d in FULL_BOOKS}
        assert exported_ids == expected_ids

    def test_iif_document_ids(self):
        text = generate_qbd_iif(FULL_BOOKS).decode("utf-8")
        for doc in FULL_BOOKS:
            assert doc["document_id"] in text, f"IIF lost document_id: {doc['document_id']}"

    def test_csv_vendor_names_exact(self):
        rows = _parse_csv_bytes(generate_csv(FULL_BOOKS), bom=True)
        exported_vendors = {r[1] for r in rows[1:]}
        expected_vendors = {d["vendor"] for d in FULL_BOOKS}
        assert exported_vendors == expected_vendors

    def test_gl_accounts_preserved(self):
        """GL account codes must not be truncated or reformatted."""
        rows = _parse_csv_bytes(generate_csv(FULL_BOOKS), bom=True)
        exported_gls = {r[3] for r in rows[1:]}
        expected_gls = {d["gl_account"] for d in FULL_BOOKS}
        assert exported_gls == expected_gls


# ===================================================================
# 8. ROUNDING — ROUND_HALF_UP to $0.01
# ===================================================================

class TestRounding:
    """All monetary values must use ROUND_HALF_UP to 0.01."""

    ROUNDING_CASES = [
        ("100.005", "100.01"),  # half-up rounds up
        ("100.004", "100.00"),  # below half rounds down
        ("100.015", "100.02"),  # banker's rounding would give .01, HALF_UP gives .02
        ("0.005", "0.01"),
        ("0.001", "0.00"),
        ("999999.995", "1000000.00"),
    ]

    @pytest.mark.parametrize("input_val,expected", ROUNDING_CASES)
    def test_round_function(self, input_val, expected):
        result = _round(Decimal(input_val))
        assert result == Decimal(expected), f"_round({input_val}) = {result}, expected {expected}"

    def test_tax_extraction_all_cents(self):
        """Every tax component must be rounded to exactly 2 decimal places."""
        for doc in FULL_BOOKS:
            amount = _dec(doc["amount"])
            taxes = _extract_taxes(amount, doc["tax_code"])
            for key in ("pre_tax", "gst", "qst", "hst"):
                val = taxes[key]
                assert val == val.quantize(CENT, rounding=ROUND_HALF_UP), (
                    f"doc={doc['document_id']} {key}={val} not rounded to cent"
                )

    def test_csv_amounts_two_decimals(self):
        """CSV amount strings must have at most 2 decimal places."""
        rows = _parse_csv_bytes(generate_csv(FULL_BOOKS), bom=True)
        for row in rows[1:]:
            for col_idx in (4, 5, 6, 7):  # Amount, GST, QST, HST
                val = row[col_idx]
                parts = val.split(".")
                if len(parts) == 2:
                    assert len(parts[1]) <= 2, f"CSV has >2 decimals: {val}"

    def test_large_amount_rounding_integrity(self):
        """999999.99 taxable must not lose precision."""
        doc = FULL_BOOKS[8]  # INV-009, $999,999.99
        taxes = _extract_taxes(_dec(doc["amount"]), doc["tax_code"])
        total = taxes["pre_tax"] + taxes["gst"] + taxes["qst"] + taxes["hst"]
        assert abs(total - Decimal("999999.99")) <= Decimal("0.02")


# ===================================================================
# 9. CROSS-FORMAT CONSISTENCY — same input, all formats agree
# ===================================================================

class TestCrossFormatConsistency:
    """All formats must agree on the same pre_tax, tax values for the same input."""

    def test_pretax_consistent_csv_sage_acomba_wave(self):
        """pre_tax amount must be identical across CSV-derived, Sage, Acomba, and Wave."""
        for doc in FULL_BOOKS:
            amount = _dec(doc["amount"])
            taxes = _extract_taxes(amount, doc["tax_code"])
            expected_pretax = taxes["pre_tax"]

            # Sage50
            sage_rows = _parse_csv_bytes(generate_sage50([doc]))
            assert Decimal(sage_rows[1][4]) == expected_pretax, (
                f"Sage50 pretax mismatch for {doc['document_id']}"
            )

            # Acomba
            acomba_rows = _parse_tsv_bytes(generate_acomba([doc]))
            assert Decimal(acomba_rows[1][4]) == expected_pretax, (
                f"Acomba pretax mismatch for {doc['document_id']}"
            )

            # Wave
            wave_rows = _parse_csv_bytes(generate_wave([doc]))
            assert Decimal(wave_rows[1][2]) == expected_pretax, (
                f"Wave pretax mismatch for {doc['document_id']}"
            )

    def test_document_count_consistent(self):
        """All formats must export the same number of documents."""
        n = len(FULL_BOOKS)
        csv_rows = _parse_csv_bytes(generate_csv(FULL_BOOKS), bom=True)
        sage_rows = _parse_csv_bytes(generate_sage50(FULL_BOOKS))
        acomba_rows = _parse_tsv_bytes(generate_acomba(FULL_BOOKS))
        xero_rows = _parse_csv_bytes(generate_xero(FULL_BOOKS))
        wave_rows = _parse_csv_bytes(generate_wave(FULL_BOOKS))

        assert len(csv_rows) - 1 == n, f"CSV: {len(csv_rows)-1} rows, expected {n}"
        assert len(sage_rows) - 1 == n, f"Sage: {len(sage_rows)-1} rows, expected {n}"
        assert len(acomba_rows) - 1 == n, f"Acomba: {len(acomba_rows)-1} rows, expected {n}"
        assert len(xero_rows) - 1 == n, f"Xero: {len(xero_rows)-1} rows, expected {n}"
        assert len(wave_rows) - 1 == n, f"Wave: {len(wave_rows)-1} rows, expected {n}"


# ===================================================================
# 10. ANNUAL ZIP — integrity across monthly boundaries
# ===================================================================

class TestAnnualZipCorruption:
    """ZIP export must not corrupt monthly totals or encoding."""

    def test_zip_monthly_csvs_decode(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        data = generate_annual_zip("CLI1", 2026, db_path)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for name in zf.namelist():
                if name.endswith(".csv"):
                    content = zf.read(name)
                    # Must decode without error
                    if content[:3] == b"\xef\xbb\xbf":
                        content[3:].decode("utf-8")
                    else:
                        content.decode("utf-8")

    def test_zip_french_survives(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        data = generate_annual_zip("CLI1", 2026, db_path)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            jan_csv = zf.read("OtoCPA_Export_CLI1_2026-01.csv")
            text = jan_csv.decode("utf-8-sig")
            assert "Verrouillé" in text, "ZIP lost French accents in locked-period CSV"

    def test_zip_no_data_duplication(self, tmp_path):
        db_path = _setup_locked_db(tmp_path)
        data = generate_annual_zip("CLI1", 2026, db_path)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            jan_csv = zf.read("OtoCPA_Export_CLI1_2026-01.csv")
            feb_csv = zf.read("OtoCPA_Export_CLI1_2026-02.csv")
            jan_text = jan_csv.decode("utf-8-sig")
            feb_text = feb_csv.decode("utf-8-sig")
            # Jan doc should only be in Jan, Feb doc only in Feb
            # Use comma-delimited match to avoid substring false positives
            assert ",LOCK-001" in jan_text or "LOCK-001," in jan_text
            assert "UNLOCK-001" not in jan_text
            assert "UNLOCK-001" in feb_text
            # Feb must not contain LOCK-001 as a standalone ID
            feb_ids = [r[9] for r in csv.reader(io.StringIO(feb_text)) if len(r) > 9]
            assert "LOCK-001" not in feb_ids
